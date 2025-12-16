import logging
from datetime import timedelta

import jwt
from django.conf import settings
from django.db import transaction
from django.utils import timezone
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Usuario, RegistrosSesion

logger = logging.getLogger(__name__)


class LoginView(APIView):
    """
    Vista para autenticación de usuarios.
    Al autenticarse correctamente, además crea un registro en RegistrosSesion.
    """
    def post(self, request):
        correo = (request.data.get('correo') or '').strip()
        contrasena = (request.data.get('contrasena') or '').strip()

        if not correo or not contrasena:
            return Response({'error': 'Correo y contraseña son requeridos'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Nota: mantengo tu verificación por correo + contraseña en texto plano
            usuario = Usuario.objects.get(correo=correo, contrasena=contrasena)
        except Usuario.DoesNotExist:
            return Response({'error': 'Credenciales inválidas'}, status=status.HTTP_401_UNAUTHORIZED)
        except Exception:
            logger.exception("Error buscando usuario")
            return Response({'error': 'Error interno'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Verifica que el estado sea activo (id_estado == 1)
        try:
            estado_id = usuario.id_estado.id_estado
        except Exception:
            estado_id = None

        if estado_id != 1:
            return Response({'error': 'usuario inactivo'}, status=status.HTTP_403_FORBIDDEN)

        # Generamos access token (2 horas) y refresh token (7 días)
        now = timezone.now()
        access_token_exp = now + timedelta(hours=2)
        refresh_token_exp = now + timedelta(days=7)

        # Payload para access token
        access_payload = {
            'id_usuario': usuario.id_usuario,
            'correo': usuario.correo,
            'type': 'access',
            'exp': int(access_token_exp.timestamp()),
            'iat': int(now.timestamp())
        }

        # Payload para refresh token
        refresh_payload = {
            'id_usuario': usuario.id_usuario,
            'correo': usuario.correo,
            'type': 'refresh',
            'exp': int(refresh_token_exp.timestamp()),
            'iat': int(now.timestamp())
        }

        try:
            access_token = jwt.encode(access_payload, settings.SECRET_KEY, algorithm='HS256')
            refresh_token = jwt.encode(refresh_payload, settings.SECRET_KEY, algorithm='HS256')
            
            # Aseguramos que sean strings
            if isinstance(access_token, bytes):
                access_token = access_token.decode('utf-8')
            if isinstance(refresh_token, bytes):
                refresh_token = refresh_token.decode('utf-8')
                
        except Exception:
            logger.exception("Error generando tokens JWT")
            return Response({'error': 'Error generando tokens'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Intentamos crear el registro de sesión (SIN refresh_token para no modificar el modelo)
        registro_data = None
        try:
            with transaction.atomic():
                id_empresa_fk = getattr(usuario, 'id_empresa', None)
                registro = RegistrosSesion.objects.create(
                    id_empresa=id_empresa_fk,
                    id_usuario=usuario,
                    fecha_inicio_sesion=timezone.now()
                    # NO incluimos refresh_token aquí para no modificar el modelo
                )
                registro.refresh_from_db()

                registro_data = {
                    'id_registro': registro.id_registro,
                    'fecha_inicio_sesion': registro.fecha_inicio_sesion.isoformat()
                }
        except Exception:
            logger.exception("No se pudo crear RegistrosSesion tras login para usuario %s", getattr(usuario, 'id_usuario', None))
            registro_data = None

        respuesta_usuario = {
            'id': usuario.id_usuario,
            'nombre': usuario.nombres,
            'correo': usuario.correo,
            'rol': getattr(usuario.id_permiso_acceso, 'rol', None),
            'estado': getattr(usuario.id_estado, 'estado', None)
        }

        return Response({
            'token': access_token,  # Mantenemos 'token' para compatibilidad
            'refresh_token': refresh_token,  # Añadimos refresh_token
            'usuario': respuesta_usuario,
            'registro_sesion': registro_data,
            'expires_in': 7200  # 2 horas en segundos
        }, status=status.HTTP_200_OK)



class RefreshTokenView(APIView):
    """
    Vista para renovar el access token usando el refresh token.
    No requiere modificar el modelo RegistrosSesion.
    """
    def post(self, request):
        refresh_token = request.data.get('refresh_token')
        
        if not refresh_token:
            return Response({'error': 'Refresh token es requerido'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Verificamos el refresh token
            payload = jwt.decode(refresh_token, settings.SECRET_KEY, algorithms=['HS256'])
            
            # Verificamos que sea un refresh token
            if payload.get('type') != 'refresh':
                return Response({'error': 'Token inválido'}, status=status.HTTP_401_UNAUTHORIZED)
            
            # Buscamos el usuario
            usuario = Usuario.objects.get(id_usuario=payload['id_usuario'])
            
            # Verificamos que el usuario esté activo
            if usuario.id_estado.id_estado != 1:
                return Response({'error': 'Usuario inactivo'}, status=status.HTTP_403_FORBIDDEN)

            # Generamos nuevo access token
            now = timezone.now()
            access_token_exp = now + timedelta(hours=2)
            
            access_payload = {
                'id_usuario': usuario.id_usuario,
                'correo': usuario.correo,
                'type': 'access',
                'exp': int(access_token_exp.timestamp()),
                'iat': int(now.timestamp())
            }
            
            access_token = jwt.encode(access_payload, settings.SECRET_KEY, algorithm='HS256')
            if isinstance(access_token, bytes):
                access_token = access_token.decode('utf-8')
                
            return Response({
                'token': access_token,  # Devolvemos como 'token' para compatibilidad
                'access_token': access_token,  # Y también como 'access_token'
                'expires_in': 7200  # 2 horas en segundos
            }, status=status.HTTP_200_OK)
            
        except jwt.ExpiredSignatureError:
            return Response({'error': 'Refresh token expirado'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'error': 'Refresh token inválido'}, status=status.HTTP_401_UNAUTHORIZED)
        except Usuario.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_401_UNAUTHORIZED)
        except Exception:
            logger.exception("Error renovando token")
            return Response({'error': 'Error interno'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

















# appdataflowai/views.py

# views.py (parche defensivo)
import jwt
import logging
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Usuario, DetalleProducto

logger = logging.getLogger(__name__)

class UsuarioInfoView(APIView):
    """
    Vista que retorna la información detallada del usuario autenticado,
    incluyendo datos personales, rol, empresa, categoría, plan, estado
    y los productos (dashboards) asignados, a partir del token JWT.
    Esta versión es defensiva: evita AttributeError por relaciones NULL
    y registra el stacktrace para debug en Render.
    """

    def get(self, request):
        # 1) Extraer token de Authorization header; si no existe -> 401
        auth_header = request.headers.get('Authorization', '')
        if not auth_header:
            logger.warning("UsuarioInfoView: Authorization header ausente")
            return Response({'error': 'Authorization header ausente'}, status=status.HTTP_401_UNAUTHORIZED)

        parts = auth_header.split()
        if len(parts) != 2:
            logger.warning("UsuarioInfoView: formato de Authorization header inválido: %s", auth_header)
            return Response({'error': 'Formato de Authorization inválido'}, status=status.HTTP_401_UNAUTHORIZED)

        token = parts[1]

        # 2) Decodificar token JWT
        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
            if not id_usuario:
                logger.warning("UsuarioInfoView: token válido pero sin id_usuario en payload")
                return Response({'error': 'Token inválido (sin id_usuario)'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.ExpiredSignatureError:
            return Response({'error': 'Token expirado'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'error': 'Token inválido'}, status=status.HTTP_401_UNAUTHORIZED)
        except Exception as e:
            logger.exception("UsuarioInfoView: error decodificando token")
            return Response({'error': 'Error al decodificar token'}, status=status.HTTP_401_UNAUTHORIZED)

        # 3) Buscar usuario y serializar defensivamente
        try:
            usuario = Usuario.objects.select_related(
                'id_empresa__id_categoria',
                'id_empresa__id_plan',
                'id_empresa__id_estado',
                'id_permiso_acceso'
            ).get(id_usuario=id_usuario)
        except Usuario.DoesNotExist:
            logger.warning("UsuarioInfoView: usuario no encontrado id_usuario=%s", id_usuario)
            return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception:
            # si hay cualquier otro error en la consulta, lo loggeamos
            logger.exception("UsuarioInfoView: error obteniendo usuario id_usuario=%s", id_usuario)
            return Response({'error': 'Error interno al obtener usuario'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            # detalles productos (si no hay nada, devolvemos lista vacía)
            detalles = DetalleProducto.objects.filter(id_usuario=usuario).select_related('id_producto')
        except Exception:
            logger.exception("UsuarioInfoView: error consultando DetalleProducto para usuario id=%s", getattr(usuario, 'id_usuario', None))
            detalles = []

        lista_productos = []
        for det in detalles:
            try:
                prod = getattr(det, 'id_producto', None)
                lista_productos.append({
                    'id_producto': getattr(prod, 'id_producto', None),
                    'producto':     getattr(prod, 'producto', None),
                    'slug':         getattr(prod, 'slug', None),
                    'iframe':       getattr(prod, 'iframe', None),
                })
            except Exception:
                logger.exception("UsuarioInfoView: error serializando un producto para usuario id=%s", getattr(usuario, 'id_usuario', None))

        # Serialización defensiva de empresa y sub-objetos
        empresa = getattr(usuario, 'id_empresa', None)
        empresa_data = None
        if empresa:
            try:
                # usar getattr por si algún campo es None o no existe
                fecha_reg = getattr(empresa, 'fecha_registros', None)
                fecha_reg_iso = fecha_reg.isoformat() if fecha_reg else None

                categoria = getattr(empresa, 'id_categoria', None)
                plan = getattr(empresa, 'id_plan', None)
                estado = getattr(empresa, 'id_estado', None)

                empresa_data = {
                    'id': getattr(empresa, 'id_empresa', None),
                    'nombre': getattr(empresa, 'nombre_empresa', None),
                    'nombre_corto': getattr(empresa, 'nombre_corto', None),
                    'direccion': getattr(empresa, 'direccion', None),
                    'fecha_registro': fecha_reg_iso,
                    'telefono': getattr(empresa, 'telefono', None),
                    'ciudad': getattr(empresa, 'ciudad', None),
                    'pais': getattr(empresa, 'pais', None),
                    'categoria': {
                        'id': getattr(categoria, 'id_categoria', None),
                        'descripcion': getattr(categoria, 'descripcion_categoria', None),
                    } if categoria else None,
                    'plan': {
                        'id': getattr(plan, 'id_plan', None),
                        'tipo': getattr(plan, 'tipo_plan', None),
                    } if plan else None,
                    'estado': {
                        'id': getattr(estado, 'id_estado', None),
                        'nombre': getattr(estado, 'estado', None),
                    } if estado else None,
                }
            except Exception:
                logger.exception("UsuarioInfoView: error serializando empresa para user_id=%s", getattr(usuario, 'id_usuario', None))
                empresa_data = None

        # Rol (puede estar en id_permiso_acceso)
        rol_obj = getattr(usuario, 'id_permiso_acceso', None)
        rol_val = getattr(rol_obj, 'rol', None) if rol_obj else None

        data = {
            'id': getattr(usuario, 'id_usuario', None),
            'nombres': getattr(usuario, 'nombres', None),
            'correo': getattr(usuario, 'correo', None),
            'rol': rol_val,
            'empresa': empresa_data,
            'productos': lista_productos
        }

        return Response(data, status=status.HTTP_200_OK)






class ProductosUsuarioView(APIView):
    """
    Vista que retorna todos los productos asociados al usuario autenticado,
    junto con información relevante del producto, área y usuario.
    """
    def get(self, request):
        token = request.headers.get('Authorization', '').split(' ')[-1]
        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
        except jwt.ExpiredSignatureError:
            return Response({'error': 'Token expirado'}, status=401)
        except jwt.InvalidTokenError:
            return Response({'error': 'Token inválido'}, status=401)

        try:
            usuario = Usuario.objects.select_related(
                'id_empresa',
                'id_permiso_acceso'
            ).get(id_usuario=id_usuario)
        except Usuario.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=404)

        detalles = DetalleProducto.objects.select_related(
            'id_producto',
            'id_producto__id_estado',
            'id_producto__id_area'
        ).filter(id_usuario=usuario)

        productos = []
        for dp in detalles:
            prod = dp.id_producto  # instancia Producto

            productos.append({
                'id_producto': getattr(prod, 'id_producto', None),
                'producto': getattr(prod, 'producto', None),
                'slug': getattr(prod, 'slug', None),
                'iframe': getattr(prod, 'iframe', None),
                'estado': getattr(prod.id_estado, 'estado', None) if getattr(prod, 'id_estado', None) else None,
                'link_pb': getattr(prod, 'link_pb', None),
                'categoria_producto': getattr(prod, 'categoria_producto', None),
                'area': {
                    'id_area': getattr(prod.id_area, 'id_area', None),
                    'nombre': getattr(prod.id_area, 'area_trabajo', None),
                },
                'link_dashboard_externo': getattr(prod, 'link_dashboard_externo', None),
                # <-- nuevo campo traído desde DetalleProducto
                'db_name': getattr(dp, 'db_name', None),
                'usuario': {
                    'id': usuario.id_usuario,
                    'nombres': usuario.nombres,
                    'correo': usuario.correo,
                    'rol': usuario.id_permiso_acceso.rol,
                    'empresa': usuario.id_empresa.nombre_empresa,
                }
            })

        return JsonResponse(productos, safe=False)




from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import get_authorization_header
from .models import Producto
from .serializers import ProductoSerializer
import jwt
from django.conf import settings

#Retornar todos los productos para el Marketplace

class ProductoListView(APIView):
    """
    Vista protegida que retorna todos los productos del sistema.
    El usuario debe estar autenticado mediante token JWT.
    """

    def get(self, request):
        # Validar el token manualmente
        auth_header = get_authorization_header(request).split()
        if not auth_header or auth_header[0].lower() != b'bearer':
            return Response({'error': 'Token no enviado'}, status=401)

        try:
            token = auth_header[1].decode('utf-8')
            jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
        except (jwt.ExpiredSignatureError, jwt.DecodeError):
            return Response({'error': 'Token inválido o expirado'}, status=401)

        # Obtener y serializar productos
        productos = Producto.objects.select_related('id_estado').all()
        serializer = ProductoSerializer(productos, many=True)
        return Response(serializer.data, status=200)
    



import jwt
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authentication import get_authorization_header

from .models import Producto, DetalleProducto, Usuario
from .serializers import DetalleProductoSerializer

# Mapa de límites por id_plan
PLAN_LIMITES = {
    1: 1,   # Basic
    2: 5,   # Professional
    3: 100,  # Enterprise (o el número que definas)
    4: 1,
    5: 5,   # Dataflow (o el número que definas)
    6:100,
}

class AdquirirDashboardView(APIView):
    """
    POST: { "id_producto": <int> }
    Crea un registro en detalle_producto si el usuario no ha excedido su límite.
    """
    def post(self, request):
        # 1) Validar y decodificar token
        auth = get_authorization_header(request).split()
        if not auth or auth[0].lower() != b'bearer':
            return Response({'error': 'Token no enviado'}, status=status.HTTP_401_UNAUTHORIZED)
        try:
            token = auth[1].decode()
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload['id_usuario']
        except jwt.ExpiredSignatureError:
            return Response({'error': 'Token expirado'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'error': 'Token inválido'}, status=status.HTTP_401_UNAUTHORIZED)

        # 2) Traer usuario y su plan
        try:
            usuario = Usuario.objects.select_related('id_empresa__id_plan').get(id_usuario=id_usuario)
        except Usuario.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        id_plan = usuario.id_empresa.id_plan.id_plan
        limite = PLAN_LIMITES.get(id_plan, 0)

        # 3) Contar dashboards ya adquiridos
        actuales = DetalleProducto.objects.filter(id_usuario=usuario).count()
        if actuales >= limite:
            return Response(
                {'error': f'Has alcanzado el límite de {limite} dashboards para tu plan.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 4) Validar que exista el producto
        id_prod = request.data.get('id_producto')
        try:
            producto = Producto.objects.get(id_producto=id_prod)
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        # 5) Crear el detalle (será único por unique_together)
        detalle = DetalleProducto.objects.create(id_usuario=usuario, id_producto=producto)
        serializer = DetalleProductoSerializer(detalle)
        return Response(serializer.data, status=status.HTTP_201_CREATED)





# backend/appdataflowai/views.py
from rest_framework import generics
from .models import Empresa, Categoria, Estado, TipoPlan, Usuario, PermisoAcceso
from .serializers import (
    EmpresaSerializer,
    UsuarioSerializer,
    CategoriaSerializer,
    EstadoSerializer,
    TipoPlanSerializer,
    PermisoAccesoSerializer
)

# Crear Empresa
class EmpresaCreateAPIView(generics.CreateAPIView):
    queryset = Empresa.objects.all()
    serializer_class = EmpresaSerializer

# Crear Usuario
class UsuarioCreateAPIView(generics.CreateAPIView):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer

# Listar Categorías
class CategoriaListAPIView(generics.ListAPIView):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer

# Listar Estados
class EstadoListAPIView(generics.ListAPIView):
    queryset = Estado.objects.all()
    serializer_class = EstadoSerializer

# Listar Planes
class TipoPlanListAPIView(generics.ListAPIView):
    queryset = TipoPlan.objects.all()
    serializer_class = TipoPlanSerializer

# Listar Permisos de Acceso
class PermisoAccesoListAPIView(generics.ListAPIView):
    queryset = PermisoAcceso.objects.all()
    serializer_class = PermisoAccesoSerializer




# backend/appdataflowai/views.py


# backend/appdataflowai/views.py

import stripe
from datetime import datetime
from django.conf import settings
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Empresa, TipoPlan, Pagos
from .serializers import CreatePaymentIntentSerializer

# Inicializa la API de Stripe con tu clave secreta
stripe.api_key = settings.STRIPE_SECRET_KEY


class CreatePaymentIntentAPIView(generics.GenericAPIView):
    """
    POST /api/create-payment-intent/
    Recibe { id_empresa, id_plan } y devuelve clientSecret para el frontend.
    """
    serializer_class = CreatePaymentIntentSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        id_empresa = serializer.validated_data['id_empresa']
        id_plan = serializer.validated_data['id_plan']

        # Obtiene empresa y plan de la base
        empresa = Empresa.objects.get(pk=id_empresa)
        plan = TipoPlan.objects.get(pk=id_plan)
        amount = int(plan.valor_plan * 100)  # en centavos

        # Crea el PaymentIntent en Stripe, añadiendo metadata para el webhook
        intent = stripe.PaymentIntent.create(
            amount=amount,
            currency='usd',
            metadata={
                'id_empresa': str(id_empresa),
                'id_plan': str(id_plan),
            },
        )
        return Response({'clientSecret': intent.client_secret})


class StripeWebhookAPIView(APIView):
    """
    POST /api/stripe-webhook/
    Endpoint público para recibir eventos de Stripe.
    """
    def post(self, request, *args, **kwargs):
        payload = request.body
        sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')
        try:
            evt = stripe.Webhook.construct_event(
                payload, sig_header, settings.STRIPE_WEBHOOK_SECRET
            )
        except Exception:
            return Response(status=400)

        # Solo nos interesa el pago exitoso
        if evt['type'] == 'payment_intent.succeeded':
            intent = evt['data']['object']
            id_empresa = int(intent['metadata']['id_empresa'])
            id_plan = int(intent['metadata']['id_plan'])
            amount = intent['amount_received'] / 100.0
            # Stripe devuelve 'created' en timestamp en segundos
            paid_at = datetime.fromtimestamp(intent['created'])

            # 1) Actualizar la fecha de pago en Empresa
            Empresa.objects.filter(pk=id_empresa).update(fecha_hora_pago=paid_at)

            # 2) Crear registro en tabla Pagos
            Pagos.objects.create(
                id_empresa_id=id_empresa,
                id_plan_id=id_plan,
                ingreso=amount,
                fecha_hora_pago=paid_at
            )

        return Response(status=200)






#DashboardVentasDataflow de PRUEBA
# appdataflowai/views.py
#
from rest_framework.permissions import IsAuthenticated
from django.utils.dateparse import parse_date
from .serializers import DashboardVentasDataflowSerializer
from .serializers import DashboardVentasSerializer


class DashboardVentasDataflowView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        usuario = request.user
        if not hasattr(usuario, 'id_usuario'):
            return Response({'error': 'Usuario inválido en request'}, status=401)

        qs = DashboardVentasDataflow.objects.filter(id_empresa=usuario.id_empresa).order_by('fecha_entrega')
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        if start:
            qs = qs.filter(fecha_entrega__gte=parse_date(start))
        if end:
            qs = qs.filter(fecha_entrega__lte=parse_date(end))

        serializer = DashboardVentasDataflowSerializer(qs, many=True)
        return Response(serializer.data, status=200)



#####
class DashboardVentasView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        usuario = request.user
        if not hasattr(usuario, 'id_usuario'):
            return Response({'error': 'Usuario inválido en request'}, status=401)

        queryset = DashboardVentas.objects.filter(id_empresa=usuario.id_empresa).order_by('fecha_venta')
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        if start:
            queryset = queryset.filter(fecha_venta__gte=parse_date(start))
        if end:
            queryset = queryset.filter(fecha_venta__lte=parse_date(end))

        serializer = DashboardVentasSerializer(queryset, many=True, context={'usuario': usuario})
        return Response(serializer.data, status=200)

    




























from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authentication import get_authorization_header
from django.utils.dateparse import parse_date
from django.conf import settings
import jwt

from .models import DashboardSales
from .serializers import DashboardSalesSerializer


class DashboardSalesView(APIView):
    """
    Protected view that returns DashboardSales records,
    with optional filtering by date range (?start=YYYY-MM-DD&end=YYYY-MM-DD)
    """

    def get(self, request):
        # 1. Manual JWT authentication
        auth_header = get_authorization_header(request).split()
        if not auth_header or auth_header[0].lower() != b'bearer':
            return Response({'error': 'Token not provided'}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            token = auth_header[1].decode('utf-8')
            jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
        except jwt.ExpiredSignatureError:
            return Response({'error': 'Token expired'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'error': 'Invalid token'}, status=status.HTTP_401_UNAUTHORIZED)

        # 2. Query and filters
        queryset = DashboardSales.objects.all().order_by('sale_date')
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        if start:
            queryset = queryset.filter(sale_date__gte=parse_date(start))
        if end:
            queryset = queryset.filter(sale_date__lte=parse_date(end))

        # 3. Serialize
        serializer = DashboardSalesSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)














# Vista para retornar registros de DashboardFinanzas
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authentication import get_authorization_header
from django.utils.dateparse import parse_date
from django.conf import settings
import jwt

from .models import DashboardFinanzas
from .serializers import DashboardFinanzasSerializer

class DashboardFinanzasView(APIView):
    """
    Vista protegida que retorna los registros de DashboardFinanzas,
    con filtrado opcional por rango de fechas (?start=YYYY-MM-DD&end=YYYY-MM-DD)
    """

    def get(self, request):
        # 1. Autenticación JWT (manual)
        auth_header = get_authorization_header(request).split()
        if not auth_header or auth_header[0].lower() != b'bearer':
            return Response({'error': 'Token no enviado'}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            token = auth_header[1].decode('utf-8')
            jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
        except jwt.ExpiredSignatureError:
            return Response({'error': 'Token expirado'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'error': 'Token inválido'}, status=status.HTTP_401_UNAUTHORIZED)

        # 2. Consulta y filtros por fecha_registro
        queryset = DashboardFinanzas.objects.all().order_by('fecha_registro')
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        if start:
            queryset = queryset.filter(fecha_registro__gte=parse_date(start))
        if end:
            queryset = queryset.filter(fecha_registro__lte=parse_date(end))

        # 3. Serializar
        serializer = DashboardFinanzasSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)





from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authentication import get_authorization_header
from django.utils.dateparse import parse_date
from django.conf import settings
import jwt

from .models import DashboardCompras
from .serializers import DashboardComprasSerializer

class DashboardComprasView(APIView):
    """
    Vista protegida que retorna los registros de DashboardCompras,
    con filtrado opcional por rango de fechas (?start=YYYY-MM-DD&end=YYYY-MM-DD)
    """

    def get(self, request):
        # 1. Autenticación JWT manual
        auth_header = get_authorization_header(request).split()
        if not auth_header or auth_header[0].lower() != b'bearer':
            return Response({'error': 'Token no enviado'}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            token = auth_header[1].decode('utf-8')
            jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
        except jwt.ExpiredSignatureError:
            return Response({'error': 'Token expirado'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'error': 'Token inválido'}, status=status.HTTP_401_UNAUTHORIZED)

        # 2. Consulta y filtros por fecha_compra
        queryset = DashboardCompras.objects.all().order_by('fecha_compra')
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        if start:
            queryset = queryset.filter(fecha_compra__gte=parse_date(start))
        if end:
            queryset = queryset.filter(fecha_compra__lte=parse_date(end))

        # 3. Serialización
        serializer = DashboardComprasSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)




# appdataflowai/views.py

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.authentication import get_authorization_header

import jwt
import pandas as pd
import logging

from django.conf import settings
from django.db import models

from .models import (
    DashboardVentasDataflow,
    DashboardVentas,
    DashboardFinanzas,
    DashboardVentasColtrade,
    DashboardVentasLoop,
    DashboardSalesreview,
    Usuario,
    Producto,
)

logger = logging.getLogger(__name__)

# Mapeo de producto-ID a modelo de dashboard
PRODUCTO_MODELO_MAP = {
    3: DashboardVentasDataflow,
    2: DashboardVentas,
    4: DashboardFinanzas,
    1: DashboardCompras,
    10: DashboardSalesreview,
    #  ...otros productos si los hubiera
}

class ImportarDatosView(APIView):
    """
    Importa un Excel y crea registros en el modelo correspondiente,
    asignando automáticamente id_empresa y id_producto desde el usuario y la URL.
    """
    parser_classes = [MultiPartParser]

    def post(self, request, id_producto):
        # 1) Autenticación JWT y extracción de usuario
        auth = get_authorization_header(request).split()
        if not auth or auth[0].lower() != b'bearer':
            return Response({'error': 'Token no enviado'}, status=status.HTTP_401_UNAUTHORIZED)
        try:
            token = auth[1].decode()
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            usuario = Usuario.objects.select_related('id_empresa').get(id_usuario=payload['id_usuario'])
        except jwt.ExpiredSignatureError:
            return Response({'error': 'Token expirado'}, status=status.HTTP_401_UNAUTHORIZED)
        except (jwt.InvalidTokenError, Usuario.DoesNotExist):
            return Response({'error': 'Token inválido o usuario no encontrado'}, status=status.HTTP_401_UNAUTHORIZED)

        # 2) Validar modelo y obtener instancia de Producto
        modelo = PRODUCTO_MODELO_MAP.get(int(id_producto))
        if not modelo:
            return Response({
                'error': f'ID de producto inválido: {id_producto}. Válidos: {list(PRODUCTO_MODELO_MAP)}'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            producto_obj = Producto.objects.get(id_producto=id_producto)
        except Producto.DoesNotExist:
            return Response({'error': f'Producto {id_producto} no encontrado'}, status=status.HTTP_400_BAD_REQUEST)

        # 3) Validar archivo Excel
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'error': 'No se proporcionó archivo'}, status=status.HTTP_400_BAD_REQUEST)
        if not archivo.name.lower().endswith(('.xlsx', '.xls')):
            return Response({'error': 'Formato no válido (.xlsx/.xls)'}, status=status.HTTP_400_BAD_REQUEST)
        if archivo.size > 10 * 1024 * 1024:
            return Response({'error': 'Archivo >10MB'}, status=status.HTTP_400_BAD_REQUEST)

        # 4) Leer y normalizar DataFrame
        try:
            df = pd.read_excel(archivo)
        except Exception as e:
            return Response({'error': f'Error leyendo Excel: {e}'}, status=status.HTTP_400_BAD_REQUEST)
        if df.empty:
            return Response({'error': 'Excel vacío'}, status=status.HTTP_400_BAD_REQUEST)

        df.columns = (
            df.columns
            .str.strip()
            .str.lower()
            .str.replace(' ', '_')
            .str.replace('-', '_')
        )

        # 5) Determinar campos del modelo EXCLUYENDO id_empresa e id_producto
        campos_modelo = [
            f.name for f in modelo._meta.fields
            if not isinstance(f, models.AutoField)
            and f.name not in ('id_empresa', 'id_producto')
        ]

        # 6) Filtrar sólo las columnas presentes en el Excel
        columnas = set(df.columns)
        campos_validos = [c for c in campos_modelo if c in columnas]
        if not campos_validos:
            return Response({
                'error': 'No hay columnas válidas',
                'esperados': campos_modelo,
                'encontrados': list(columnas)
            }, status=status.HTTP_400_BAD_REQUEST)

        # 7) Procesar fila por fila
        creados = 0
        errores = []
        for idx, row in df.iterrows():
            datos = {}
            try:
                # Rellenar campos válidos
                for campo in campos_validos:
                    valor = row[campo]
                    if pd.isna(valor):
                        datos[campo] = None
                    else:
                        field = modelo._meta.get_field(campo)
                        if isinstance(field, models.DateField):
                            datos[campo] = pd.to_datetime(valor).date()
                        elif isinstance(field, models.TimeField):
                            datos[campo] = pd.to_datetime(valor).time()
                        elif isinstance(field, models.IntegerField):
                            datos[campo] = int(valor)
                        elif isinstance(field, models.DecimalField):
                            datos[campo] = float(valor)
                        else:
                            datos[campo] = valor

                # Asignar foráneas obligatorias desde contexto
                datos['id_empresa'] = usuario.id_empresa
                datos['id_producto'] = producto_obj

                # Crear registro
                modelo.objects.create(**datos)
                creados += 1

            except Exception as e:
                errores.append(f'Fila {idx+2}: {e}')
                if len(errores) > 50:
                    errores.append('Detenido por exceso de errores')
                    break

        # 8) Responder
        resp = {
            'mensaje': f'{creados} registros creados correctamente',
            'totales': len(df),
            'creados': creados,
            'campos_importados': campos_validos,
        }
        if errores:
            resp['errores'] = errores[:10]
            resp['total_errores'] = len(errores)

        return Response(resp, status=status.HTTP_200_OK)


class EstadoImportacionView(APIView):
    """
    Retorna estadísticas básicas del modelo de importación.
    """
    def get(self, request, id_producto):
        auth = get_authorization_header(request).split()
        if not auth or auth[0].lower() != b'bearer':
            return Response({'error': 'Token no enviado'}, status=status.HTTP_401_UNAUTHORIZED)
        try:
            token = auth[1].decode()
            jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
        except jwt.ExpiredSignatureError:
            return Response({'error': 'Token expirado'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'error': 'Token inválido'}, status=status.HTTP_401_UNAUTHORIZED)

        modelo = PRODUCTO_MODELO_MAP.get(int(id_producto))
        if not modelo:
            return Response({'error': 'Producto inválido'}, status=status.HTTP_400_BAD_REQUEST)

        total = modelo.objects.count()
        campos = [
            f.name for f in modelo._meta.fields
            if not isinstance(f, models.AutoField)
        ]
        return Response({
            'id_producto': id_producto,
            'modelo': modelo.__name__,
            'total_registros': total,
            'campos': campos,
            'tabla': modelo._meta.db_table
        }, status=status.HTTP_200_OK)




"""

Endpoints para página de perfil

"""
# your_app/views.py
import jwt
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import AuthenticationFailed
from .models import Usuario
from .serializers import PasswordChangeSerializer

class CambiarContrasenaView(APIView):
    """
    PATCH: Cambia la contraseña del usuario autenticado por JWT.
    Body JSON:
    {
      "contrasena_actual": "laActual",
      "contrasena_nueva": "laNueva"
    }
    (ATENCIÓN: guarda la contraseña en texto plano tal cual se recibe)
    """

    def _get_usuario_from_token(self, request):
        auth = request.headers.get('Authorization', '')
        token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else auth.split(' ')[-1] if auth else ''
        if not token:
            raise AuthenticationFailed('No se proporcionó token')

        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
            if not id_usuario:
                raise AuthenticationFailed('Token inválido')
        except jwt.ExpiredSignatureError:
            raise AuthenticationFailed('Token expirado')
        except jwt.InvalidTokenError:
            raise AuthenticationFailed('Token inválido')

        try:
            usuario = Usuario.objects.get(id_usuario=id_usuario)
            return usuario
        except Usuario.DoesNotExist:
            raise AuthenticationFailed('Usuario no encontrado')

    def patch(self, request):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        serializer = PasswordChangeSerializer(data=request.data, context={'user': usuario})
        if serializer.is_valid():
            serializer.save()
            return Response({'detail': 'Contraseña actualizada correctamente.'}, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)




#modificar usuario y empresa
# views.py
import jwt
from django.conf import settings
from django.db import IntegrityError, transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, serializers
from rest_framework.exceptions import AuthenticationFailed

from .models import Usuario, Empresa, Estado, TipoPlan, PermisoAcceso


def _get_usuario_from_token(request):
    """
    Extrae y valida el token JWT del header Authorization y devuelve la instancia Usuario.
    Lanza AuthenticationFailed si algo falla.
    """
    auth = request.headers.get('Authorization', '')
    token = ''
    if auth:
        token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else auth.split(' ')[-1]

    if not token:
        raise AuthenticationFailed('No se proporcionó token')

    try:
        payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
        id_usuario = payload.get('id_usuario')
        if not id_usuario:
            raise AuthenticationFailed('Token inválido')
    except jwt.ExpiredSignatureError:
        raise AuthenticationFailed('Token expirado')
    except jwt.InvalidTokenError:
        raise AuthenticationFailed('Token inválido')

    try:
        usuario = Usuario.objects.select_related('id_empresa__id_plan', 'id_estado', 'id_permiso_acceso').get(id_usuario=id_usuario)
        return usuario
    except Usuario.DoesNotExist:
        raise AuthenticationFailed('Usuario no encontrado')


# Serializers para actualización

class UsuarioUpdateSerializer(serializers.Serializer):
    # Campos permitidos para editar en Usuario (NO se incluyen los campos bloqueados)
    nombres = serializers.CharField(max_length=200, required=False)
    apellidos = serializers.CharField(max_length=200, allow_blank=True, required=False)
    correo = serializers.EmailField(required=False)

    def validate_correo(self, value):
        # La vista debe pasar el usuario actual en context para comparar
        usuario_actual = self.context.get('usuario_actual')
        if usuario_actual and usuario_actual.correo == value:
            return value
        # Verificar unicidad
        if Usuario.objects.filter(correo=value).exists():
            raise serializers.ValidationError('El correo ya está en uso')
        return value


class EmpresaUpdateSerializer(serializers.Serializer):
    # Campos permitidos para editar en Empresa (los campos bloqueados NO están aquí)
    nombre_empresa = serializers.CharField(max_length=200, required=False)
    direccion = serializers.CharField(max_length=200, required=False)
    telefono = serializers.CharField(max_length=20, required=False)
    ciudad = serializers.CharField(max_length=100, required=False)
    pais = serializers.CharField(max_length=100, required=False)
    prefijo_pais = serializers.CharField(max_length=5, allow_blank=True, required=False)
    correo = serializers.EmailField(required=False, allow_null=True)
    pagina_web = serializers.URLField(required=False, allow_null=True)

    # validaciones adicionales si se requieren se pueden agregar aquí


class PerfilMeView(APIView):
    """
    GET: devuelve info del usuario autenticado y la empresa asociada (lectura).
    PATCH: actualiza la información del usuario autenticado (solo campos permitidos).
    Ruta: GET/PATCH /perfil/me/
    """
    def get(self, request):
        try:
            usuario = _get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = usuario.id_empresa

        resp = {
            'usuario': {
                'id_usuario': usuario.id_usuario,
                'nombres': usuario.nombres,
                'apellidos': usuario.apellidos,
                'correo': usuario.correo,
                'id_permiso_acceso': getattr(usuario.id_permiso_acceso, 'id_permiso_acceso', None),
                'rol': getattr(usuario.id_permiso_acceso, 'rol', None),
                'id_estado': getattr(usuario.id_estado, 'id_estado', None),
                'estado': getattr(usuario.id_estado, 'estado', None),
            },
            'empresa': {
                'id_empresa': empresa.id_empresa,
                'id_plan': getattr(empresa.id_plan, 'id_plan', None),
                'nombre_empresa': empresa.nombre_empresa,
                'direccion': empresa.direccion,
                'telefono': empresa.telefono,
                'ciudad': empresa.ciudad,
                'pais': empresa.pais,
                'prefijo_pais': empresa.prefijo_pais,
                'correo': empresa.correo,
                'pagina_web': empresa.pagina_web,
                'id_estado': getattr(empresa.id_estado, 'id_estado', None),
                'fecha_registros': empresa.fecha_registros.isoformat() if empresa.fecha_registros else None,
                'fecha_hora_pago': empresa.fecha_hora_pago.isoformat() if empresa.fecha_hora_pago else None,
            }
        }
        return Response(resp, status=status.HTTP_200_OK)

    def patch(self, request):
        """
        Actualiza los campos permitidos del usuario autenticado.
        NO permite actualizar: id_usuario, id_empresa, id_permiso_acceso, id_estado, contrasena.
        """
        try:
            usuario = _get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        serializer = UsuarioUpdateSerializer(data=request.data, context={'usuario_actual': usuario})
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        # Actualizar
        try:
            with transaction.atomic():
                if 'nombres' in data:
                    usuario.nombres = data['nombres']
                if 'apellidos' in data:
                    usuario.apellidos = data['apellidos']
                if 'correo' in data:
                    usuario.correo = data['correo']
                usuario.save()
        except IntegrityError:
            return Response({'error': 'No se pudo actualizar usuario. Posible conflicto de datos.'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'detail': 'Usuario actualizado correctamente',
            'usuario': {
                'id_usuario': usuario.id_usuario,
                'nombres': usuario.nombres,
                'apellidos': usuario.apellidos,
                'correo': usuario.correo
            }
        }, status=status.HTTP_200_OK)


class EmpresaView(APIView):
    """
    GET: devuelve los datos de la empresa del usuario autenticado.
    PATCH: actualiza los campos permitidos de la empresa (solo si el usuario es admin: id_permiso_acceso == 1)
    Ruta: GET/PATCH /perfil/empresa/
    """
    def get(self, request):
        try:
            usuario = _get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = usuario.id_empresa
        resp = {
            'empresa': {
                'id_empresa': empresa.id_empresa,
                'id_plan': getattr(empresa.id_plan, 'id_plan', None),
                'nombre_empresa': empresa.nombre_empresa,
                'direccion': empresa.direccion,
                'telefono': empresa.telefono,
                'ciudad': empresa.ciudad,
                'pais': empresa.pais,
                'prefijo_pais': empresa.prefijo_pais,
                'correo': empresa.correo,
                'pagina_web': empresa.pagina_web,
                'id_estado': getattr(empresa.id_estado, 'id_estado', None),
                'fecha_registros': empresa.fecha_registros.isoformat() if empresa.fecha_registros else None,
                'fecha_hora_pago': empresa.fecha_hora_pago.isoformat() if empresa.fecha_hora_pago else None,
            }
        }
        return Response(resp, status=status.HTTP_200_OK)

    def patch(self, request):
        """
        Actualizar empresa: SOLO si el usuario autenticado tiene id_permiso_acceso == 1
        Campos NO editables: id_empresa, id_plan, id_estado, fecha_registros, fecha_hora_pago
        Campos editables: nombre_empresa, direccion, telefono, ciudad, pais, prefijo_pais, correo, pagina_web
        """
        try:
            usuario = _get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        # comprobación de permiso (solo rol con id_permiso_acceso == 1 puede editar empresa)
        permiso_id = getattr(usuario.id_permiso_acceso, 'id_permiso_acceso', None)
        if permiso_id != 1:
            return Response({'error': 'No autorizado. Se requiere permiso administrativo para editar la empresa.'}, status=status.HTTP_403_FORBIDDEN)

        empresa = usuario.id_empresa

        serializer = EmpresaUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        try:
            with transaction.atomic():
                if 'nombre_empresa' in data:
                    empresa.nombre_empresa = data['nombre_empresa']
                if 'direccion' in data:
                    empresa.direccion = data['direccion']
                if 'telefono' in data:
                    empresa.telefono = data['telefono']
                if 'ciudad' in data:
                    empresa.ciudad = data['ciudad']
                if 'pais' in data:
                    empresa.pais = data['pais']
                if 'prefijo_pais' in data:
                    empresa.prefijo_pais = data['prefijo_pais']
                if 'correo' in data:
                    empresa.correo = data['correo']
                if 'pagina_web' in data:
                    empresa.pagina_web = data['pagina_web']
                empresa.save()
        except IntegrityError:
            return Response({'error': 'No se pudo actualizar la empresa. Posible conflicto de datos.'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'detail': 'Empresa actualizada correctamente',
            'empresa': {
                'id_empresa': empresa.id_empresa,
                'nombre_empresa': empresa.nombre_empresa,
                'direccion': empresa.direccion,
                'telefono': empresa.telefono,
                'ciudad': empresa.ciudad,
                'pais': empresa.pais,
                'prefijo_pais': empresa.prefijo_pais,
                'correo': empresa.correo,
                'pagina_web': empresa.pagina_web,
            }
        }, status=status.HTTP_200_OK)















#ACTIVAR O DESACTIVAR USUARIOS
# views.py
import jwt
import logging
from django.conf import settings
from django.db import IntegrityError, transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, serializers
from rest_framework.exceptions import AuthenticationFailed

from .models import Usuario, Empresa, Estado, TipoPlan, PermisoAcceso, Areas

logger = logging.getLogger(__name__)

# Map de límites por plan: None = ilimitado
PLAN_LIMITES = {
    1: 1,    # Basic
    2: 3,    # Professional
    3: 100,  # Enterprise -> ilimitado
    4: 1,
    5: 3,
    6: 100,  # Unlimited
}


def _get_usuario_from_token(request):
    auth = request.headers.get('Authorization', '')
    token = ''
    if auth:
        token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else auth.split(' ')[-1]

    if not token:
        raise AuthenticationFailed('No se proporcionó token')

    try:
        payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
        id_usuario = payload.get('id_usuario')
        if not id_usuario:
            raise AuthenticationFailed('Token inválido')
    except jwt.ExpiredSignatureError:
        raise AuthenticationFailed('Token expirado')
    except jwt.InvalidTokenError:
        raise AuthenticationFailed('Token inválido')

    try:
        usuario = Usuario.objects.select_related('id_empresa__id_plan', 'id_estado', 'id_permiso_acceso', 'id_area').get(id_usuario=id_usuario)
        return usuario
    except Usuario.DoesNotExist:
        raise AuthenticationFailed('Usuario no encontrado')


# Serializers

class EstadoChangeSerializer(serializers.Serializer):
    id_estado = serializers.IntegerField()


class CreateUserSerializer(serializers.Serializer):
    nombres = serializers.CharField(max_length=200)
    apellidos = serializers.CharField(max_length=200, allow_blank=True, required=False)
    correo = serializers.EmailField()
    contrasena = serializers.CharField(max_length=255)
    contrasena_confirm = serializers.CharField(max_length=255, write_only=True)
    id_permiso_acceso = serializers.IntegerField(required=False)
    id_estado = serializers.IntegerField(required=False)
    id_area = serializers.IntegerField(required=True)  # ahora obligatorio desde frontend

    def validate_id_permiso_acceso(self, value):
        try:
            PermisoAcceso.objects.get(id_permiso_acceso=value)
        except PermisoAcceso.DoesNotExist:
            raise serializers.ValidationError('PermisoAcceso no existe')
        return value

    def validate_id_estado(self, value):
        try:
            Estado.objects.get(id_estado=value)
        except Estado.DoesNotExist:
            raise serializers.ValidationError('Estado no existe')
        return value

    def validate_id_area(self, value):
        try:
            Areas.objects.get(id_area=value)
        except Areas.DoesNotExist:
            raise serializers.ValidationError('Área no existe')
        return value

    def validate(self, data):
        if data.get('contrasena') != data.get('contrasena_confirm'):
            raise serializers.ValidationError({'contrasena_confirm': 'Las contraseñas no coinciden'})

        correo = data.get('correo', '').strip().lower()
        data['correo'] = correo
        return data


class RoleChangeSerializer(serializers.Serializer):
    id_permiso_acceso = serializers.IntegerField()

    def validate_id_permiso_acceso(self, value):
        try:
            PermisoAcceso.objects.get(id_permiso_acceso=value)
        except PermisoAcceso.DoesNotExist:
            raise serializers.ValidationError('PermisoAcceso no existe')
        return value


class UsuariosEmpresaView(APIView):
    """
    GET: lista usuarios de la misma empresa del usuario autenticado.
    POST: crea un usuario en la misma empresa (respectando limites del plan).
    Ruta: /perfil/usuarios/
    """
    def get(self, request):
        try:
            usuario = _get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = usuario.id_empresa
        usuarios_qs = Usuario.objects.filter(id_empresa=empresa).select_related('id_estado', 'id_permiso_acceso', 'id_area')

        usuarios = []
        for u in usuarios_qs:
            usuarios.append({
                'id_usuario': u.id_usuario,
                'nombres': u.nombres,
                'apellidos': u.apellidos,
                'correo': u.correo,
                'id_permiso_acceso': getattr(u.id_permiso_acceso, 'id_permiso_acceso', None),
                'rol': getattr(u.id_permiso_acceso, 'rol', None),
                'id_estado': getattr(u.id_estado, 'id_estado', None),
                'estado': getattr(u.id_estado, 'estado', None),
                'id_area': getattr(u.id_area, 'id_area', None),
                'area_trabajo': getattr(u.id_area, 'area_trabajo', None),
            })

        return Response({'usuarios': usuarios}, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            usuario_aut = _get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        serializer = CreateUserSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        empresa = usuario_aut.id_empresa
        try:
            plan_id = empresa.id_plan.id_plan
        except Exception:
            plan_id = None
        limite = PLAN_LIMITES.get(plan_id, None)

        usuarios_count = Usuario.objects.filter(id_empresa=empresa).count()
        if limite is not None and usuarios_count >= limite:
            return Response({
                'error': 'Límite de usuarios alcanzado para el plan de la empresa',
                'limit': limite,
                'current': usuarios_count
            }, status=status.HTTP_403_FORBIDDEN)

        correo_normalizado = data['correo']
        if Usuario.objects.filter(correo=correo_normalizado).exists():
            return Response({'error': 'Correo ya registrado'}, status=status.HTTP_400_BAD_REQUEST)

        id_permiso_acceso = data.get('id_permiso_acceso', None)
        if id_permiso_acceso is None:
            id_permiso_acceso = getattr(usuario_aut.id_permiso_acceso, 'id_permiso_acceso', None)

        try:
            permiso_obj = PermisoAcceso.objects.get(id_permiso_acceso=id_permiso_acceso)
        except PermisoAcceso.DoesNotExist:
            permiso_obj = usuario_aut.id_permiso_acceso

        id_estado = data.get('id_estado', None)
        if id_estado is None:
            id_estado = 1
        try:
            estado_obj = Estado.objects.get(id_estado=id_estado)
        except Estado.DoesNotExist:
            return Response({'error': 'Estado no existe'}, status=status.HTTP_400_BAD_REQUEST)

        # Validar area (ahora obligatorio en serializer)
        id_area = data.get('id_area')
        try:
            area_obj = Areas.objects.get(id_area=id_area)
        except Areas.DoesNotExist:
            return Response({'error': 'Área no existe'}, status=status.HTTP_400_BAD_REQUEST)

        # Crear usuario
        try:
            with transaction.atomic():
                nuevo = Usuario.objects.create(
                    id_empresa=empresa,
                    id_permiso_acceso=permiso_obj,
                    id_area=area_obj,
                    nombres=data['nombres'],
                    apellidos=data.get('apellidos', '') or '',
                    correo=correo_normalizado,
                    contrasena=data['contrasena'],
                    id_estado=estado_obj
                )
        except IntegrityError as e:
            msg = str(e)
            logger.exception("IntegrityError creando usuario: %s", e)
            if 'unique' in msg.lower() or 'duplicate' in msg.lower() or 'correo' in msg.lower():
                return Response({'error': 'Correo ya registrado (unique constraint).'}, status=status.HTTP_400_BAD_REQUEST)
            return Response({'error': 'No se pudo crear usuario. Error de integridad en la base de datos.'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.exception("Error creando usuario: %s", e)
            return Response({'error': 'Error interno al crear usuario.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            'detail': 'Usuario creado correctamente',
            'usuario': {
                'id_usuario': nuevo.id_usuario,
                'nombres': nuevo.nombres,
                'apellidos': nuevo.apellidos,
                'correo': nuevo.correo,
                'id_estado': nuevo.id_estado.id_estado,
                'estado': nuevo.id_estado.estado,
                'id_area': nuevo.id_area.id_area,
                'area_trabajo': nuevo.id_area.area_trabajo
            }
        }, status=status.HTTP_201_CREATED)


class UsuarioEstadoChangeView(APIView):
    def patch(self, request, id_usuario):
        try:
            usuario_autenticado = _get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            usuario_obj = Usuario.objects.select_related('id_empresa', 'id_estado').get(id_usuario=id_usuario)
        except Usuario.DoesNotExist:
            return Response({'error': 'Usuario objetivo no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        if usuario_obj.id_empresa.id_empresa != usuario_autenticado.id_empresa.id_empresa:
            return Response({'error': 'No autorizado para modificar usuarios de otra empresa'}, status=status.HTTP_403_FORBIDDEN)

        serializer = EstadoChangeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        nuevo_estado_id = serializer.validated_data['id_estado']
        try:
            nuevo_estado = Estado.objects.get(id_estado=nuevo_estado_id)
        except Estado.DoesNotExist:
            return Response({'error': 'Estado no existe'}, status=status.HTTP_400_BAD_REQUEST)

        usuario_obj.id_estado = nuevo_estado
        usuario_obj.save()

        return Response({
            'detail': 'Estado actualizado correctamente',
            'id_usuario': usuario_obj.id_usuario,
            'id_estado': nuevo_estado.id_estado,
            'estado': nuevo_estado.estado
        }, status=status.HTTP_200_OK)


class UsuarioRolChangeView(APIView):
    def patch(self, request, id_usuario):
        try:
            usuario_autenticado = _get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        permiso_act = getattr(usuario_autenticado.id_permiso_acceso, 'id_permiso_acceso', None)
        if permiso_act != 1:
            return Response({'error': 'No autorizado. Se requiere permiso administrativo para cambiar roles.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            usuario_obj = Usuario.objects.select_related('id_empresa', 'id_permiso_acceso').get(id_usuario=id_usuario)
        except Usuario.DoesNotExist:
            return Response({'error': 'Usuario objetivo no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        if usuario_obj.id_empresa.id_empresa != usuario_autenticado.id_empresa.id_empresa:
            return Response({'error': 'No autorizado para modificar usuarios de otra empresa'}, status=status.HTTP_403_FORBIDDEN)

        if usuario_obj.id_usuario == usuario_autenticado.id_usuario:
            return Response({'error': 'No puedes cambiar tu propio rol'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = RoleChangeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        nuevo_rol_id = serializer.validated_data['id_permiso_acceso']
        try:
            nuevo_rol = PermisoAcceso.objects.get(id_permiso_acceso=nuevo_rol_id)
        except PermisoAcceso.DoesNotExist:
            return Response({'error': 'PermisoAcceso no existe'}, status=status.HTTP_400_BAD_REQUEST)

        usuario_obj.id_permiso_acceso = nuevo_rol
        usuario_obj.save()

        return Response({
            'detail': 'Rol actualizado correctamente',
            'id_usuario': usuario_obj.id_usuario,
            'id_permiso_acceso': nuevo_rol.id_permiso_acceso,
            'rol': nuevo_rol.rol
        }, status=status.HTTP_200_OK)


class UsuarioDeleteView(APIView):
    def delete(self, request, id_usuario):
        try:
            usuario_autenticado = _get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            usuario_obj = Usuario.objects.select_related('id_empresa').get(id_usuario=id_usuario)
        except Usuario.DoesNotExist:
            return Response({'error': 'Usuario objetivo no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        if usuario_obj.id_empresa.id_empresa != usuario_autenticado.id_empresa.id_empresa:
            return Response({'error': 'No autorizado para eliminar usuarios de otra empresa'}, status=status.HTTP_403_FORBIDDEN)

        if usuario_obj.id_usuario == usuario_autenticado.id_usuario:
            return Response({'error': 'No puedes eliminar tu propio usuario'}, status=status.HTTP_400_BAD_REQUEST)

        usuario_obj.delete()
        return Response({'detail': 'Usuario eliminado correctamente'}, status=status.HTTP_204_NO_CONTENT)


class PermisosListView(APIView):
    def get(self, request):
        permisos_qs = PermisoAcceso.objects.all()
        permisos = []
        for p in permisos_qs:
            permisos.append({
                'id_permiso_acceso': p.id_permiso_acceso,
                'rol': p.rol
            })
        return Response({'permisos': permisos}, status=status.HTTP_200_OK)


class AreasListView(APIView):
    """
    GET: lista todas las áreas.
    Ruta: GET /perfil/areas/
    """
    def get(self, request):
        areas_qs = Areas.objects.all()
        areas = [{'id_area': a.id_area, 'area_trabajo': a.area_trabajo} for a in areas_qs]
        return Response({'areas': areas}, status=status.HTTP_200_OK)


















































#ASOCIAR DASHBOARDS POR MEDIO DE PERFIL
# views.py
import jwt
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import AuthenticationFailed
from django.db import transaction

from .models import Usuario, Producto, DetalleProducto, EmpresaDashboard, DashboardContext
from .serializers import (
    AsgDashboardUsuarioListSerializer,
    AsgDashboardProductoSerializer,
    AsgDashboardDetalleProductoSerializer,
    DashboardContextSerializer
)

# Mapa de límites por id_plan (usa el tuyo)
PLAN_LIMITES = {
    1: 1,   # Basic
    2: 5,   # Professional
    3: 1000,  # Enterprise
    4: 1,
    5: 5,   # Dataflow
    6: 1000,
}


class AsgDashboardBasePerfilView(APIView):
    """
    Clase base para obtener usuario desde token.
    """

    def asgdashboard_get_usuario_from_token(self, request):
        auth = request.headers.get('Authorization', '')
        token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else auth.split(' ')[-1] if auth else ''
        if not token:
            raise AuthenticationFailed('No se proporcionó token')

        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
            if not id_usuario:
                raise AuthenticationFailed('Token inválido')
        except jwt.ExpiredSignatureError:
            raise AuthenticationFailed('Token expirado')
        except jwt.InvalidTokenError:
            raise AuthenticationFailed('Token inválido')

        try:
            usuario = Usuario.objects.get(id_usuario=id_usuario)
            return usuario
        except Usuario.DoesNotExist:
            raise AuthenticationFailed('Usuario no encontrado')


class AsgDashboardUsuariosEmpresaView(AsgDashboardBasePerfilView):
    """
    GET /asg/perfil/usuarios/  -> Lista los usuarios que pertenecen a la misma empresa
    """
    def get(self, request):
        try:
            usuario = self.asgdashboard_get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        usuarios = Usuario.objects.filter(id_empresa=usuario.id_empresa)
        serializer = AsgDashboardUsuarioListSerializer(usuarios, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class AsgDashboardProductosListView(AsgDashboardBasePerfilView):
    """
    GET /asg/perfil/productos/  -> Lista de productos disponibles para la empresa del usuario que hace la petición.
    Regla: se excluyen productos que estén asociados (en EmpresaDashboard) a otras empresas.
    """
    def get(self, request):
        try:
            requester = self.asgdashboard_get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        # Productos que están asociados a OTRAS empresas
        productos_bloqueados_ids = EmpresaDashboard.objects.exclude(empresa=requester.id_empresa).values_list('producto_id', flat=True)
        productos = Producto.objects.exclude(id_producto__in=productos_bloqueados_ids).order_by('producto')
        serializer = AsgDashboardProductoSerializer(productos, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class AsgDashboardUsuarioAsignacionesView(AsgDashboardBasePerfilView):
    """
    GET  /asg/perfil/usuarios/<id_usuario>/asignaciones/  -> devuelve productos asignados a ese usuario
    POST /asg/perfil/usuarios/<id_usuario>/asignaciones/  -> asigna un producto a ese usuario (body: { "id_producto": <int> })
    """

    def asgdashboard_get_target_usuario_or_401(self, request, id_usuario):
        try:
            requester = self.asgdashboard_get_usuario_from_token(request)
        except AuthenticationFailed as e:
            raise AuthenticationFailed(str(e))
        try:
            target = Usuario.objects.get(id_usuario=id_usuario)
        except Usuario.DoesNotExist:
            return None, Response({'error': 'Usuario objetivo no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        # Solo permitir operar sobre usuarios de la misma empresa
        if target.id_empresa.id_empresa != requester.id_empresa.id_empresa:
            return None, Response({'error': 'No autorizado para ver/editar usuarios de otra empresa.'}, status=status.HTTP_403_FORBIDDEN)

        return target, None

    def get(self, request, id_usuario):
        try:
            target, err_resp = self.asgdashboard_get_target_usuario_or_401(request, id_usuario)
            if err_resp:
                return err_resp
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        asignaciones = DetalleProducto.objects.filter(id_usuario=target)
        serializer = AsgDashboardDetalleProductoSerializer(asignaciones, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @transaction.atomic
    def post(self, request, id_usuario):
        """
        Body: { "id_producto": <int> }
        Al asignar:
          - se crea DetalleProducto
          - se crea (o recupera) DashboardContext con session_id='Dataflow' y los campos tomados del Producto
          - se devuelve el detalle serializado (incluye dashboard_context via serializer)
        """
        try:
            target, err_resp = self.asgdashboard_get_target_usuario_or_401(request, id_usuario)
            if err_resp:
                return err_resp
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        id_producto = request.data.get('id_producto')
        if not id_producto:
            return Response({'error': 'Se requiere id_producto'}, status=status.HTTP_400_BAD_REQUEST)

        # Verificar existencia del producto
        try:
            producto = Producto.objects.get(id_producto=id_producto)
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        # ** NUEVA VALIDACIÓN: si el producto está asociado en EmpresaDashboard a OTRA empresa -> bloquear **
        pertenece_otra_empresa = EmpresaDashboard.objects.filter(producto=producto).exclude(empresa=target.id_empresa).exists()
        if pertenece_otra_empresa:
            return Response({'error': 'Producto pertenece a otra empresa y no puede asignarse.'}, status=status.HTTP_400_BAD_REQUEST)

        # Límite por plan de la empresa del usuario objetivo
        plan_obj = target.id_empresa.id_plan  # objeto TipoPlan
        plan_id = getattr(plan_obj, 'id_plan', None)
        limite = PLAN_LIMITES.get(plan_id, None)
        if limite is None:
            # sin límite por defecto (puedes ajustar)
            limite = None

        # Conteo actual de asignaciones del usuario objetivo
        cuenta_actual = DetalleProducto.objects.filter(id_usuario=target).count()
        if limite is not None and cuenta_actual >= limite:
            return Response({
                'error': 'Límite de dashboards alcanzado para este usuario según su plan.',
                'limite': limite,
                'asignados': cuenta_actual
            }, status=status.HTTP_400_BAD_REQUEST)

        # Verificar si ya existe la asociación
        exists = DetalleProducto.objects.filter(id_usuario=target, id_producto=producto).exists()
        if exists:
            return Response({'error': 'Producto ya asignado a este usuario.'}, status=status.HTTP_400_BAD_REQUEST)

        # Crear la asociación
        detalle = DetalleProducto.objects.create(id_usuario=target, id_producto=producto)

        # --- Crear o recuperar DashboardContext asociado a esta empresa+nombre ---
        # Obtener empresa_id como entero
        empresa_fk = getattr(target, 'id_empresa', None)
        empresa_id = None
        if empresa_fk is not None:
            empresa_id = getattr(empresa_fk, 'id_empresa', empresa_fk)  # si FK o int

        dc_defaults = {
            'dashboard_context': producto.dashboard_context or '',
            'tables': producto.tables or {},
            'formularios_id': producto.formularios_id or None,
        }

        # Usamos get_or_create para evitar duplicados por empresa+nombre
        dc, created = DashboardContext.objects.get_or_create(
            session_id='Dataflow',
            dashboard_name=producto.producto,
            empresa_id=empresa_id,
            defaults=dc_defaults
        )
        # Si ya existía y deseas actualizarlo con la info del producto, descomenta:
        # if not created:
        #     dc.dashboard_context = producto.dashboard_context or dc.dashboard_context
        #     dc.tables = producto.tables or dc.tables
        #     dc.formularios_id = producto.formularios_id or dc.formularios_id
        #     dc.save()

        # Serializar detalle (incluye dashboard_context via serializer)
        serializer = AsgDashboardDetalleProductoSerializer(detalle)
        # También devolvemos info resumida del DashboardContext creado/recuperado
        dc_ser = DashboardContextSerializer(dc).data if dc else None
        response_data = serializer.data
        response_data['_dashboard_context'] = {
            'created': bool(created),
            'dashboard_context': dc_ser
        }
        return Response(response_data, status=status.HTTP_201_CREATED)


class AsgDashboardUsuarioEliminarAsignacionView(AsgDashboardBasePerfilView):
    """
    DELETE /asg/perfil/usuarios/<id_usuario>/asignaciones/<id_producto>/  -> elimina la asignación (opcional)
    """
    def delete(self, request, id_usuario, id_producto):
        try:
            requester = self.asgdashboard_get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            target = Usuario.objects.get(id_usuario=id_usuario)
        except Usuario.DoesNotExist:
            return Response({'error': 'Usuario objetivo no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        if target.id_empresa.id_empresa != requester.id_empresa.id_empresa:
            return Response({'error': 'No autorizado.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            producto = Producto.objects.get(id_producto=id_producto)
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        deleted, _ = DetalleProducto.objects.filter(id_usuario=target, id_producto=producto).delete()
        if deleted:
            return Response({'detail': 'Asignación eliminada.'}, status=status.HTTP_204_NO_CONTENT)
        else:
            return Response({'error': 'Asignación no encontrada.'}, status=status.HTTP_404_NOT_FOUND)






















































# Vista para retornar registros de DashboardSalesreview
from .models import DashboardSalesreview
from .serializers import DashboardSalesreviewSerializer

class DashboardSalesreviewView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        usuario = request.user
        if not hasattr(usuario, 'id_usuario'):
            return Response({'error': 'Usuario inválido en request'}, status=status.HTTP_401_UNAUTHORIZED)
        try:
            queryset = DashboardSalesreview.objects.filter(id_empresa=usuario.id_empresa).order_by('fecha_compra')
        except Exception as exc:
            return Response({'error': 'Error al consultar datos'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        if start:
            date_start = parse_date(start)
            if not date_start:
                return Response({'error': 'Formato de fecha "start" inválido. Use YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)
            queryset = queryset.filter(fecha_compra__gte=date_start)
        if end:
            date_end = parse_date(end)
            if not date_end:
                return Response({'error': 'Formato de fecha "end" inválido. Use YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)
            queryset = queryset.filter(fecha_compra__lte=date_end)

        serializer = DashboardSalesreviewSerializer(queryset, many=True, context={'usuario': usuario})
        return Response(serializer.data, status=status.HTTP_200_OK)







#VISTA PARA SOPORTE DE MODULO
import jwt
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import AuthenticationFailed
from .models import Ticket, Usuario
from .serializers import TicketSerializer

class TokenHelperMixin:
    """
    Mixin con el método para obtener el usuario desde el token (misma lógica que tu ejemplo).
    """
    def _get_usuario_from_token(self, request):
        auth = request.headers.get('Authorization', '')
        token = ''
        if 'Bearer ' in auth:
            token = auth.split('Bearer ')[-1]
        elif auth:
            token = auth.split(' ')[-1]
        if not token:
            raise AuthenticationFailed('No se proporcionó token')

        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
            if not id_usuario:
                raise AuthenticationFailed('Token inválido')
        except jwt.ExpiredSignatureError:
            raise AuthenticationFailed('Token expirado')
        except jwt.InvalidTokenError:
            raise AuthenticationFailed('Token inválido')

        try:
            usuario = Usuario.objects.get(id_usuario=id_usuario)
            return usuario
        except Usuario.DoesNotExist:
            raise AuthenticationFailed('Usuario no encontrado')

class TicketListCreateView(TokenHelperMixin, APIView):
    """
    GET: Lista todos los tickets del usuario autenticado.
    POST: Crea un ticket asignado al usuario del token.
    """
    def get(self, request):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        tickets = Ticket.objects.filter(id_usuario=usuario).order_by('-fecha_creacion')
        serializer = TicketSerializer(tickets, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        # Tomamos solo los campos que el usuario debe enviar
        data = {
            'correo': request.data.get('correo'),
            'asunto': request.data.get('asunto'),
            'descripcion': request.data.get('descripcion', ''),
            # comentario lo dejamos vacío por ahora; estado se asigna por default en el modelo
        }

        serializer = TicketSerializer(data=data)
        if serializer.is_valid():
            # crear instancia pero asegurando id_usuario
            ticket = Ticket(
                id_usuario=usuario,
                correo=serializer.validated_data['correo'],
                asunto=serializer.validated_data['asunto'],
                descripcion=serializer.validated_data.get('descripcion', '')
            )
            ticket.save()
            out_serialized = TicketSerializer(ticket)
            return Response(out_serialized.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class TicketDetailView(TokenHelperMixin, APIView):
    """
    GET: Detalle de un ticket (el ticket debe pertenecer al usuario autenticado).
    """
    def get(self, request, pk):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            ticket = Ticket.objects.get(pk=pk)
        except Ticket.DoesNotExist:
            return Response({'error': 'Ticket no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        if ticket.id_usuario != usuario:
            return Response({'error': 'No autorizado para ver este ticket'}, status=status.HTTP_403_FORBIDDEN)

        serializer = TicketSerializer(ticket)
        return Response(serializer.data, status=status.HTTP_200_OK)





#VISTA PARA RETORNAR LAS HERRAMIENTAS CORRESPONDIENTES A CADA USUARIO
import jwt
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Usuario, DetalleProductoHerramientas


class HerramientasUsuarioView(APIView):
    """
    Retorna una lista de productos (producto_herramienta) asociados al usuario autenticado,
    sin información del usuario.
    """
    def get(self, request):
        token = request.headers.get('Authorization', '').split(' ')[-1]
        if not token:
            return Response({'error': 'Token no proporcionado'}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
        except jwt.ExpiredSignatureError:
            return Response({'error': 'Token expirado'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'error': 'Token inválido'}, status=status.HTTP_401_UNAUTHORIZED)

        # traemos los detalles vinculados al usuario (evita N+1)
        detalles = DetalleProductoHerramientas.objects.select_related(
            'id_producto',
            'id_producto__id_area',
            'id_producto__id_estado'
        ).filter(id_usuario__id_usuario=id_usuario)

        productos = []
        for dp in detalles:
            prod = dp.id_producto  # instancia ProductoHerramientas

            # nombre del campo del estado/área puede variar; try varios nombres comunes
            estado_obj = getattr(prod, 'id_estado', None)
            estado_val = None
            if estado_obj is not None:
                estado_val = getattr(estado_obj, 'estado', None) or getattr(estado_obj, 'nombre', None)

            area_obj = getattr(prod, 'id_area', None)
            area_id = getattr(area_obj, 'id_area', None) or getattr(area_obj, 'pk', None)
            area_nombre = getattr(area_obj, 'area_trabajo', None) or getattr(area_obj, 'nombre', None)

            productos.append({
                'id_producto': getattr(prod, 'id_producto_herramienta', None) or getattr(prod, 'pk', None),
                'producto': getattr(prod, 'producto_herramienta', None),
                'tipo_producto': getattr(prod, 'tipo_producto', None),
                'link_producto': getattr(prod, 'link_producto', None),
                'estado': estado_val,
                'area': {
                    'id_area': area_id,
                    'nombre': area_nombre,
                }
            })

        return Response(productos, status=status.HTTP_200_OK)




#
#Vista para el crud del dashboard de SALESREVIEW del modelo: DashboardSalesreview
# views.py
import jwt
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import AuthenticationFailed
from .models import DashboardSalesreview, Usuario
from .serializers import DashboardSalesreviewSerializer
from django.shortcuts import get_object_or_404
from datetime import datetime

class DashboardSalesreviewListCreate(APIView):
    """
    GET (lista con filtros opcionales): ?mes=abril  OR ?mes_numero=4 OR ?fecha_from=2025-04-01&fecha_to=2025-04-30
    POST: crea un registro (id_producto se forzará a DEFAULT_PRODUCT_ID en el serializer).
    """

    def _get_usuario_from_token(self, request):
        auth = request.headers.get('Authorization', '')
        token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else auth.split(' ')[-1] if auth else ''
        if not token:
            raise AuthenticationFailed('No se proporcionó token')

        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
            if not id_usuario:
                raise AuthenticationFailed('Token inválido')
        except jwt.ExpiredSignatureError:
            raise AuthenticationFailed('Token expirado')
        except jwt.InvalidTokenError:
            raise AuthenticationFailed('Token inválido')

        try:
            usuario = Usuario.objects.select_related('id_empresa').get(id_usuario=id_usuario)
            return usuario
        except Usuario.DoesNotExist:
            raise AuthenticationFailed('Usuario no encontrado')

    def _apply_filters(self, qs, request):
        """
        Aplica filtros de query params al queryset.
        Soporta:
        - mes (nombre, case-insensitive, exact match)
        - mes_numero (int)
        - fecha_from (YYYY-MM-DD)
        - fecha_to (YYYY-MM-DD)
        """
        q = qs
        mes = request.query_params.get('mes')
        mes_numero = request.query_params.get('mes_numero')
        fecha_from = request.query_params.get('fecha_from')
        fecha_to = request.query_params.get('fecha_to')

        if mes:
            q = q.filter(mes__iexact=mes)
        if mes_numero:
            try:
                q = q.filter(mes_numero=int(mes_numero))
            except ValueError:
                pass
        if fecha_from:
            try:
                dfrom = datetime.strptime(fecha_from, '%Y-%m-%d').date()
                q = q.filter(fecha_compra__gte=dfrom)
            except ValueError:
                pass
        if fecha_to:
            try:
                dto = datetime.strptime(fecha_to, '%Y-%m-%d').date()
                q = q.filter(fecha_compra__lte=dto)
            except ValueError:
                pass
        return q

    def get(self, request):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        qs = DashboardSalesreview.objects.filter(id_empresa=empresa)
        qs = self._apply_filters(qs, request)
        serializer = DashboardSalesreviewSerializer(qs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = DashboardSalesreviewSerializer(data=request.data, context={'empresa': empresa})
        if serializer.is_valid():
            obj = serializer.save()
            return Response(DashboardSalesreviewSerializer(obj).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DashboardSalesreviewDetail(APIView):
    """
    GET/PUT/PATCH/DELETE por pk. Mantiene la validación de empresa.
    """

    def _get_usuario_from_token(self, request):
        auth = request.headers.get('Authorization', '')
        token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else auth.split(' ')[-1] if auth else ''
        if not token:
            raise AuthenticationFailed('No se proporcionó token')

        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
            if not id_usuario:
                raise AuthenticationFailed('Token inválido')
        except jwt.ExpiredSignatureError:
            raise AuthenticationFailed('Token expirado')
        except jwt.InvalidTokenError:
            raise AuthenticationFailed('Token inválido')

        try:
            usuario = Usuario.objects.select_related('id_empresa').get(id_usuario=id_usuario)
            return usuario
        except Usuario.DoesNotExist:
            raise AuthenticationFailed('Usuario no encontrado')

    def get(self, request, pk):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        obj = get_object_or_404(DashboardSalesreview, pk=pk)
        if obj.id_empresa != empresa:
            return Response({'error': 'No autorizado para ver este registro'}, status=status.HTTP_403_FORBIDDEN)

        serializer = DashboardSalesreviewSerializer(obj)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def _update(self, request, pk, partial):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        obj = get_object_or_404(DashboardSalesreview, pk=pk)
        if obj.id_empresa != empresa:
            return Response({'error': 'No autorizado para modificar este registro'}, status=status.HTTP_403_FORBIDDEN)

        data = dict(request.data)
        data.pop('id_empresa', None)
        data.pop('id_producto', None)
        serializer = DashboardSalesreviewSerializer(obj, data=data, partial=partial, context={'empresa': empresa})
        if serializer.is_valid():
            updated = serializer.save()
            return Response(DashboardSalesreviewSerializer(updated).data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def delete(self, request, pk):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        obj = get_object_or_404(DashboardSalesreview, pk=pk)
        if obj.id_empresa != empresa:
            return Response({'error': 'No autorizado para eliminar este registro'}, status=status.HTTP_403_FORBIDDEN)

        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class DashboardSalesreviewBulkDelete(APIView):
    """
    POST: borra en masa los registros que coinciden con los filtros y pertenecen a la empresa del usuario.
    Body (JSON) o query params soportados (se usan los query params si se llaman desde frontend sin body):
    - mes (nombre)
    - mes_numero (int)
    - fecha_from (YYYY-MM-DD)
    - fecha_to (YYYY-MM-DD)

    Responde: {'deleted': <cantidad>}
    """

    def _get_usuario_from_token(self, request):
        auth = request.headers.get('Authorization', '')
        token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else auth.split(' ')[-1] if auth else ''
        if not token:
            raise AuthenticationFailed('No se proporcionó token')

        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
            if not id_usuario:
                raise AuthenticationFailed('Token inválido')
        except jwt.ExpiredSignatureError:
            raise AuthenticationFailed('Token expirado')
        except jwt.InvalidTokenError:
            raise AuthenticationFailed('Token inválido')

        try:
            usuario = Usuario.objects.select_related('id_empresa').get(id_usuario=id_usuario)
            return usuario
        except Usuario.DoesNotExist:
            raise AuthenticationFailed('Usuario no encontrado')

    def post(self, request):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        # Preparamos filtros: permitimos recibir por body (JSON) o por query params
        params = {}
        if request.data:
            params = request.data
        else:
            params = request.query_params

        # Construimos queryset base
        qs = DashboardSalesreview.objects.filter(id_empresa=empresa)
        # Re-utilizamos la lógica de filtrado (copia simple)
        mes = params.get('mes')
        mes_numero = params.get('mes_numero')
        fecha_from = params.get('fecha_from')
        fecha_to = params.get('fecha_to')

        if mes:
            qs = qs.filter(mes__iexact=mes)
        if mes_numero:
            try:
                qs = qs.filter(mes_numero=int(mes_numero))
            except ValueError:
                pass
        if fecha_from:
            try:
                dfrom = datetime.strptime(fecha_from, '%Y-%m-%d').date()
                qs = qs.filter(fecha_compra__gte=dfrom)
            except ValueError:
                pass
        if fecha_to:
            try:
                dto = datetime.strptime(fecha_to, '%Y-%m-%d').date()
                qs = qs.filter(fecha_compra__lte=dto)
            except ValueError:
                pass

        count = qs.count()
        if count == 0:
            return Response({'deleted': 0, 'detail': 'No se encontraron registros para eliminar'}, status=status.HTTP_200_OK)

        qs.delete()
        return Response({'deleted': count}, status=status.HTTP_200_OK)


from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class DashboardSalesreviewExport(APIView):
    """
    GET: Exporta los registros filtrados de la empresa del usuario como un archivo Excel (.xlsx).
    Soporta los mismos filtros que DashboardSalesreviewListCreate:
    - mes
    - mes_numero
    - fecha_from (YYYY-MM-DD)
    - fecha_to   (YYYY-MM-DD)
    """

    def _get_usuario_from_token(self, request):
        auth = request.headers.get('Authorization', '')
        token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else auth.split(' ')[-1] if auth else ''
        if not token:
            raise AuthenticationFailed('No se proporcionó token')

        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
            if not id_usuario:
                raise AuthenticationFailed('Token inválido')
        except jwt.ExpiredSignatureError:
            raise AuthenticationFailed('Token expirado')
        except jwt.InvalidTokenError:
            raise AuthenticationFailed('Token inválido')

        try:
            usuario = Usuario.objects.select_related('id_empresa').get(id_usuario=id_usuario)
            return usuario
        except Usuario.DoesNotExist:
            raise AuthenticationFailed('Usuario no encontrado')

    def _apply_filters(self, qs, request):
        """
        Copia simple de la lógica de filtros usada en ListCreate.
        """
        q = qs
        mes = request.query_params.get('mes')
        mes_numero = request.query_params.get('mes_numero')
        fecha_from = request.query_params.get('fecha_from')
        fecha_to = request.query_params.get('fecha_to')

        if mes:
            q = q.filter(mes__iexact=mes)
        if mes_numero:
            try:
                q = q.filter(mes_numero=int(mes_numero))
            except ValueError:
                pass
        if fecha_from:
            try:
                dfrom = datetime.strptime(fecha_from, '%Y-%m-%d').date()
                q = q.filter(fecha_compra__gte=dfrom)
            except ValueError:
                pass
        if fecha_to:
            try:
                dto = datetime.strptime(fecha_to, '%Y-%m-%d').date()
                q = q.filter(fecha_compra__lte=dto)
            except ValueError:
                pass
        return q

    def get(self, request):
        # autenticar
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        # obtener queryset y aplicar filtros
        qs = DashboardSalesreview.objects.filter(id_empresa=empresa)
        qs = self._apply_filters(qs, request)

        # serializar
        serializer = DashboardSalesreviewSerializer(qs, many=True)
        data = serializer.data  # lista de OrderedDicts

        # Preparar workbook con openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "DashboardSalesReview"

        # Determinar headers: si no hay registros usar fields del serializer
        if data and len(data) > 0:
            headers = list(data[0].keys())
        else:
            # si no hay datos, tomamos los campos definidos en el serializer Meta
            try:
                headers = list(DashboardSalesreviewSerializer.Meta.fields)
            except Exception:
                headers = [
                    'id_registro','id','mes','mes_numero','semana','dia_compra','fecha_compra','fecha_envio',
                    'numero_pedido','numero_oc','estado','linea','fuente','sku_enviado','categoria','producto',
                    'precio_unidad_antes_iva','unidades','ingresos_antes_iva'
                ]

        # escribir headers
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)

        # escribir filas
        for row_idx, row in enumerate(data, start=2):
            for col_idx, h in enumerate(headers, start=1):
                val = row.get(h, None)
                # normalizar tipos: las fechas serializadas ya vienen como strings (YYYY-MM-DD) desde el serializer
                ws.cell(row=row_idx, column=col_idx, value=val)

        # opcional: autoajustar anchos de columna (básico)
        for i, _ in enumerate(headers, start=1):
            col = get_column_letter(i)
            max_length = 0
            for cell in ws[col]:
                try:
                    if cell.value:
                        s = str(cell.value)
                        if len(s) > max_length:
                            max_length = len(s)
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[col].width = adjusted_width

        # guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # nombre de archivo (incluye id_empresa o pk)
        empresa_id = getattr(empresa, 'id_empresa', None) or getattr(empresa, 'pk', None) or 'empresa'
        today = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"dashboard_salesreview_{empresa_id}_{today}.xlsx"

        # respuesta HttpResponse con attachment
        from django.http import HttpResponse
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    


# vista para SHOPIFY 
#Otro comentario
# views.py
import re
import time
import requests
from django.conf import settings
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.views.decorators.csrf import csrf_exempt

# --- estilo Stripe: tomamos la config desde settings (que ya usa decouple) ---
SHOPIFY_SHOP_DOMAIN = getattr(settings, "SHOPIFY_SHOP_DOMAIN", None)
SHOPIFY_ACCESS_TOKEN = getattr(settings, "SHOPIFY_ACCESS_TOKEN", None)
API_VERSION = getattr(settings, "SHOPIFY_API_VERSION", "2025-10")

# Crear headers a nivel de módulo (igual que stripe.api_key = settings.STRIPE_SECRET_KEY)
SHOPIFY_HEADERS = {
    "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
    "Content-Type": "application/json",
}

def _fetch_from_shopify(resource_path, shop=None, token=None, limit=50, fetch_all=False, timeout=20):
    """Helper compacto para obtener recursos (ej. products.json)."""
    shop = shop or SHOPIFY_SHOP_DOMAIN
    token = token or SHOPIFY_ACCESS_TOKEN

    if not shop or not token:
        # Mismo comportamiento simple que en tu ejemplo: falla en la llamada si no hay token.
        raise requests.RequestException(
            "Falta configuración: define SHOPIFY_SHOP_DOMAIN y SHOPIFY_ACCESS_TOKEN en tu .env/settings"
        )

    headers = {
        "X-Shopify-Access-Token": token,
        "Content-Type": "application/json",
    }

    base_url = f"https://{shop}/admin/api/{API_VERSION}/{resource_path}"
    params = {"limit": max(1, min(250, int(limit)))}

    results = []
    url = base_url

    while True:
        resp = requests.get(url, headers=headers, params=params if url == base_url else None, timeout=timeout)

        # retry sencillo si rate-limited
        if resp.status_code == 429:
            time.sleep(2)
            resp = requests.get(url, headers=headers, params=params if url == base_url else None, timeout=timeout)

        resp.raise_for_status()
        data = resp.json()

        key = resource_path.split(".")[0]  # 'products'
        items = data.get(key, [])

        if not fetch_all:
            return items

        results.extend(items)

        link = resp.headers.get("Link", "")
        m = re.search(r'<([^>]+)>;\s*rel="next"', link)
        if m:
            url = m.group(1)
            params = None
            continue
        break

    return results

@csrf_exempt
@require_GET
def shopify_products(request):
    """
    GET /shopify/products/?limit=50
    Opcional: ?all=1 para traer TODAS las páginas (paginación automática).
    Opcional: ?shop=otra-tienda.myshopify.com para sobrescribir el shop configurado.
    """
    try:
        limit = int(request.GET.get("limit", 50))
    except (TypeError, ValueError):
        return JsonResponse({"error": "limit debe ser entero"}, status=400)

    fetch_all = request.GET.get("all") in ("1", "true", "True")
    shop = request.GET.get("shop") or SHOPIFY_SHOP_DOMAIN

    if not SHOPIFY_ACCESS_TOKEN or not shop:
        return JsonResponse(
            {"error": "Falta configuración de SHOPIFY_ACCESS_TOKEN o SHOPIFY_SHOP_DOMAIN en settings/.env"},
            status=500
        )

    try:
        products = _fetch_from_shopify("products.json", shop=shop, limit=limit, fetch_all=fetch_all)
    except requests.RequestException as e:
        return JsonResponse({"error": "Error al consultar Shopify", "detail": str(e)}, status=500)

    return JsonResponse({"products": products}, json_dumps_params={"ensure_ascii": False})







#odoo
# backend/app/views.py
import xmlrpc.client
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authentication import get_authorization_header
from django.conf import settings
import jwt

def safe_get_name(field_value):
    """Retorna el name de un many2one Odoo (lista [id, name]) o el valor en crudo."""
    if not field_value:
        return None
    if isinstance(field_value, (list, tuple)) and len(field_value) >= 2:
        return field_value[1]
    return field_value

class OdooSales2025View(APIView):
    """
    GET: devuelve ventas (sale.order) del año configurado (por defecto 2025) con state='sale'.
    Autenticación: Authorization: Bearer <jwt>

    Query params:
      - page (int, default 1)
      - per_page (int, default 100, max 1000)
      - all=1 to traer TODO (ignora paginación; usar con cuidado)
      - year=YYYY (opcional para sobrescribir ODOO_SALES_YEAR)
    """
    def get(self, request):
        # --- JWT auth (igual que antes) ---
        auth_header = get_authorization_header(request).split()
        if not auth_header or auth_header[0].lower() != b'bearer':
            return Response({'error': 'Token no enviado'}, status=status.HTTP_401_UNAUTHORIZED)
        try:
            token = auth_header[1].decode('utf-8')
            jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
        except jwt.ExpiredSignatureError:
            return Response({'error': 'Token expirado'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.InvalidTokenError:
            return Response({'error': 'Token inválido'}, status=status.HTTP_401_UNAUTHORIZED)
        except Exception as e:
            return Response({'error': f'Error validando token: {str(e)}'}, status=status.HTTP_401_UNAUTHORIZED)

        # --- Leer configuración Odoo desde settings (viene del .env) ---
        ODOO_URL = getattr(settings, "ODOO_URL", None)
        ODOO_DB = getattr(settings, "ODOO_DB", None)
        ODOO_USERNAME = getattr(settings, "ODOO_USERNAME", None)
        ODOO_API_KEY = getattr(settings, "ODOO_API_KEY", None)
        YEAR = request.query_params.get("year") or getattr(settings, "ODOO_SALES_YEAR", 2025)
        try:
            YEAR = int(YEAR)
        except Exception:
            YEAR = 2025

        # Validación mínima de config
        if not (ODOO_URL and ODOO_DB and ODOO_USERNAME and ODOO_API_KEY):
            return Response({'error': 'Falta configuración ODOO (revisa variables en settings/.env)'},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # --- Conexión a Odoo ---
        try:
            common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common", allow_none=True)
            uid = common.authenticate(ODOO_DB, ODOO_USERNAME, ODOO_API_KEY, {})
            if not uid:
                return Response({'error': 'Autenticación Odoo fallida'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object", allow_none=True)
        except Exception as e:
            return Response({'error': f'Error conectando a Odoo: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # --- Parámetros de paginación / opción all ---
        try:
            page = int(request.query_params.get('page', 1))
            per_page = int(request.query_params.get('per_page', 100))
        except Exception:
            page = 1
            per_page = 100
        if per_page <= 0:
            per_page = 100
        per_page = min(per_page, 1000)
        fetch_all = request.query_params.get('all') in ('1', 'true', 'True')

        # --- Dominio para año y state='sale' ---
        start_dt = f"{YEAR}-01-01 00:00:00"
        end_dt = f"{YEAR}-12-31 23:59:59"
        domain = [
            ("create_date", ">=", start_dt),
            ("create_date", "<=", end_dt),
            ("state", "=", "sale"),
        ]

        try:
            # 1) obtener ids (solo IDs, operación rápida)
            sale_order_ids = models.execute_kw(
                ODOO_DB, uid, ODOO_API_KEY,
                "sale.order", "search",
                [domain],
                {"order": "create_date desc"}
            )
        except Exception as e:
            return Response({'error': f'Error buscando sale.order ids: {str(e)}'},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        total = len(sale_order_ids)

        if total == 0:
            return Response({
                "total": 0,
                "page": page,
                "per_page": per_page,
                "results": []
            }, status=status.HTTP_200_OK)

        # Paginación sobre la lista de ids (rápido)
        if fetch_all:
            page_ids = sale_order_ids
            page = 1
            per_page = total
        else:
            start = (page - 1) * per_page
            end = start + per_page
            page_ids = sale_order_ids[start:end]

        # 2) leer registros de sale.order en batch (reduce llamadas)
        sale_fields = [
            "id",
            "name",
            "create_date",
            "x_studio_canal",
            "x_studio_orden_fuente",
            "x_studio_fuente_1",
            "commitment_date",
            "partner_id",
            "user_id",
            "company_id",
            "amount_total",
            "state",
            "invoice_ids",
            "invoice_status",
            "cart_quantity",
            "order_line",
        ]
        try:
            sale_records = models.execute_kw(
                ODOO_DB, uid, ODOO_API_KEY,
                "sale.order", "read",
                [page_ids],
                {"fields": sale_fields}
            )
        except Exception as e:
            return Response({'error': f'Error leyendo sale.order: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # 3) Colectar order_line ids y invoice ids para lecturas en batch
        all_order_line_ids = []
        all_invoice_ids = []
        for s in sale_records:
            ol = s.get("order_line") or []
            all_order_line_ids.extend(ol)
            inv = s.get("invoice_ids") or []
            all_invoice_ids.extend(inv)

        all_order_line_ids = list(set(all_order_line_ids))
        all_invoice_ids = list(set(all_invoice_ids))

        # 4) Leer order lines en batch
        order_line_map = {}
        if all_order_line_ids:
            try:
                ol_fields = ["id", "order_id", "product_id", "product_uom_qty", "price_unit", "price_subtotal", "qty_invoiced", "qty_delivered"]
                order_lines = models.execute_kw(
                    ODOO_DB, uid, ODOO_API_KEY,
                    "sale.order.line", "read",
                    [all_order_line_ids],
                    {"fields": ol_fields}
                )
                for ln in order_lines:
                    order_line_map[ln["id"]] = ln
            except Exception:
                order_line_map = {}

        # 5) Leer productos en batch para obtener marcas
        product_ids = set()
        for ln in order_line_map.values():
            p = ln.get("product_id")
            if p and isinstance(p, (list, tuple)):
                product_ids.add(p[0])
        product_brand_by_id = {}
        if product_ids:
            try:
                prod_fields = ["id", "x_studio_marca", "product_tmpl_id", "name"]
                prod_recs = models.execute_kw(
                    ODOO_DB, uid, ODOO_API_KEY,
                    "product.product", "read",
                    [list(product_ids)],
                    {"fields": prod_fields}
                )
                tmpl_ids = []
                pid_to_tmpl = {}
                for p in prod_recs:
                    pid = p.get("id")
                    if p.get("x_studio_marca"):
                        product_brand_by_id[pid] = p.get("x_studio_marca")
                    else:
                        pt = p.get("product_tmpl_id")
                        if pt and isinstance(pt, (list, tuple)):
                            pid_to_tmpl[pid] = pt[0]
                            tmpl_ids.append(pt[0])
                if tmpl_ids:
                    tmpl_records = models.execute_kw(
                        ODOO_DB, uid, ODOO_API_KEY,
                        "product.template", "read",
                        [list(set(tmpl_ids))],
                        {"fields": ["id", "x_studio_marca"]}
                    )
                    tmpl_map = {t["id"]: t.get("x_studio_marca") for t in tmpl_records}
                    for pid, tid in pid_to_tmpl.items():
                        if tmpl_map.get(tid):
                            product_brand_by_id[pid] = tmpl_map.get(tid)
            except Exception:
                product_brand_by_id = {}

        # 6) Leer facturas en batch si existen
        invoices_map = {}
        if all_invoice_ids:
            try:
                inv_fields = ["id", "name", "state", "amount_total", "invoice_date"]
                inv_recs = models.execute_kw(
                    ODOO_DB, uid, ODOO_API_KEY,
                    "account.move", "read",
                    [all_invoice_ids],
                    {"fields": inv_fields}
                )
                for inv in inv_recs:
                    invoices_map[inv["id"]] = inv
            except Exception:
                invoices_map = {}

        # 7) Armar resultados compactos (pero completos) por sale.order
        results = []
        for s in sale_records:
            order_lines_out = []
            for ln_id in s.get("order_line") or []:
                ln = order_line_map.get(ln_id)
                if not ln:
                    continue
                p = ln.get("product_id")
                prod_id = p[0] if p and isinstance(p, (list, tuple)) else None
                prod_name = p[1] if p and isinstance(p, (list, tuple)) else None
                order_lines_out.append({
                    "id": ln.get("id"),
                    "product": {
                        "id": prod_id,
                        "name": prod_name,
                        "brand": product_brand_by_id.get(prod_id, None)
                    },
                    "product_uom_qty": ln.get("product_uom_qty"),
                    "price_unit": ln.get("price_unit"),
                    "price_subtotal": ln.get("price_subtotal"),
                    "qty_invoiced": ln.get("qty_invoiced"),
                    "qty_delivered": ln.get("qty_delivered"),
                })

            invoices_detail = []
            for inv_id in s.get("invoice_ids") or []:
                inv = invoices_map.get(inv_id)
                if inv:
                    invoices_detail.append({
                        "id": inv.get("id"),
                        "name": inv.get("name"),
                        "state": inv.get("state"),
                        "amount_total": inv.get("amount_total"),
                        "invoice_date": inv.get("invoice_date"),
                    })

            results.append({
                "id": s.get("id"),
                "reference": s.get("name"),
                "create_date": s.get("create_date"),
                "canal": s.get("x_studio_canal"),
                "orden_fuente": s.get("x_studio_orden_fuente"),
                "fuente": s.get("x_studio_fuente_1"),
                "commitment_date": s.get("commitment_date"),
                "partner": {
                    "id": s.get("partner_id")[0] if s.get("partner_id") and isinstance(s.get("partner_id"), (list, tuple)) else None,
                    "name": safe_get_name(s.get("partner_id")),
                },
                "user": {
                    "id": s.get("user_id")[0] if s.get("user_id") and isinstance(s.get("user_id"), (list, tuple)) else None,
                    "name": safe_get_name(s.get("user_id")),
                },
                "company": {
                    "id": s.get("company_id")[0] if s.get("company_id") and isinstance(s.get("company_id"), (list, tuple)) else None,
                    "name": safe_get_name(s.get("company_id")),
                },
                "amount_total": s.get("amount_total"),
                "state": s.get("state"),
                "invoice_ids": s.get("invoice_ids", []),
                "invoice_status": s.get("invoice_status"),
                "cart_quantity": s.get("cart_quantity"),
                "order_lines": order_lines_out,
                "invoices_detail": invoices_detail,
            })

        # 8) Respuesta
        return Response({
            "total": total,
            "page": page,
            "per_page": per_page,
            "results": results
        }, status=status.HTTP_200_OK)



























#Vista para el crud del dashboard de SALES CORPORATIVO del modelo: DashboardSalesCorporativo
# views.py
import jwt
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import AuthenticationFailed
from django.shortcuts import get_object_or_404
from datetime import datetime

from .models import DashboardSalesCorporativo, Usuario
from .serializers import DashboardSalesCorporativoSerializerProd15

# Funcion comun para obtener usuario desde token (nombre con prod15 para evitar colisiones)
def get_usuario_from_token_prod15(request):
    auth = request.headers.get('Authorization', '')
    token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else auth.split(' ')[-1] if auth else ''
    if not token:
        raise AuthenticationFailed('No se proporciono token')
    try:
        payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
        id_usuario = payload.get('id_usuario')
        if not id_usuario:
            raise AuthenticationFailed('Token invalido')
    except jwt.ExpiredSignatureError:
        raise AuthenticationFailed('Token expirado')
    except jwt.InvalidTokenError:
        raise AuthenticationFailed('Token invalido')
    try:
        usuario = Usuario.objects.select_related('id_empresa').get(id_usuario=id_usuario)
        return usuario
    except Usuario.DoesNotExist:
        raise AuthenticationFailed('Usuario no encontrado')

class DashboardSalesCorporativoListCreateProd15(APIView):
    """
    GET: lista con filtros opcionales (mes_nombre, fecha_from, fecha_to, nombre_cliente, marca)
    POST: crea un registro (id_producto forzado a DEFAULT_PRODUCT_ID en el serializer)
    """

    def _apply_filters_prod15(self, qs, request):
        q = qs
        mes_nombre = request.query_params.get('mes_nombre')
        nombre_cliente = request.query_params.get('nombre_cliente')
        marca = request.query_params.get('marca')
        fecha_from = request.query_params.get('fecha_from')
        fecha_to = request.query_params.get('fecha_to')

        if mes_nombre:
            q = q.filter(mes_nombre__iexact=mes_nombre)
        if nombre_cliente:
            q = q.filter(nombre_cliente__icontains=nombre_cliente)
        if marca:
            q = q.filter(marca__icontains=marca)
        if fecha_from:
            try:
                dfrom = datetime.strptime(fecha_from, '%Y-%m-%d').date()
                q = q.filter(fecha__gte=dfrom)
            except ValueError:
                pass
        if fecha_to:
            try:
                dto = datetime.strptime(fecha_to, '%Y-%m-%d').date()
                q = q.filter(fecha__lte=dto)
            except ValueError:
                pass
        return q

    def get(self, request):
        try:
            usuario = get_usuario_from_token_prod15(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        qs = DashboardSalesCorporativo.objects.filter(id_empresa=empresa)
        qs = self._apply_filters_prod15(qs, request)
        serializer = DashboardSalesCorporativoSerializerProd15(qs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            usuario = get_usuario_from_token_prod15(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = DashboardSalesCorporativoSerializerProd15(data=request.data, context={'empresa': empresa})
        if serializer.is_valid():
            obj = serializer.save()
            return Response(DashboardSalesCorporativoSerializerProd15(obj).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DashboardSalesCorporativoDetailProd15(APIView):
    """
    GET/PUT/PATCH/DELETE por pk. Valida que el registro pertenezca a la misma empresa del usuario.
    """

    def _update_prod15(self, request, pk, partial):
        try:
            usuario = get_usuario_from_token_prod15(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        obj = get_object_or_404(DashboardSalesCorporativo, pk=pk)
        if obj.id_empresa != empresa:
            return Response({'error': 'No autorizado para modificar este registro'}, status=status.HTTP_403_FORBIDDEN)

        data = dict(request.data)
        data.pop('id_empresa', None)
        data.pop('id_producto', None)
        serializer = DashboardSalesCorporativoSerializerProd15(obj, data=data, partial=partial, context={'empresa': empresa})
        if serializer.is_valid():
            updated = serializer.save()
            return Response(DashboardSalesCorporativoSerializerProd15(updated).data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request, pk):
        try:
            usuario = get_usuario_from_token_prod15(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        obj = get_object_or_404(DashboardSalesCorporativo, pk=pk)
        if obj.id_empresa != empresa:
            return Response({'error': 'No autorizado para ver este registro'}, status=status.HTTP_403_FORBIDDEN)

        serializer = DashboardSalesCorporativoSerializerProd15(obj)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, pk):
        return self._update_prod15(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update_prod15(request, pk, partial=True)

    def delete(self, request, pk):
        try:
            usuario = get_usuario_from_token_prod15(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        obj = get_object_or_404(DashboardSalesCorporativo, pk=pk)
        if obj.id_empresa != empresa:
            return Response({'error': 'No autorizado para eliminar este registro'}, status=status.HTTP_403_FORBIDDEN)

        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class DashboardSalesCorporativoBulkDeleteProd15(APIView):
    """
    POST: borra registros por filtros y empresa del usuario. Responde {'deleted': <cantidad>}
    """

    def post(self, request):
        try:
            usuario = get_usuario_from_token_prod15(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        params = request.data if request.data else request.query_params

        qs = DashboardSalesCorporativo.objects.filter(id_empresa=empresa)
        mes_nombre = params.get('mes_nombre')
        nombre_cliente = params.get('nombre_cliente')
        marca = params.get('marca')
        fecha_from = params.get('fecha_from')
        fecha_to = params.get('fecha_to')

        if mes_nombre:
            qs = qs.filter(mes_nombre__iexact=mes_nombre)
        if nombre_cliente:
            qs = qs.filter(nombre_cliente__icontains=nombre_cliente)
        if marca:
            qs = qs.filter(marca__icontains=marca)
        if fecha_from:
            try:
                dfrom = datetime.strptime(fecha_from, '%Y-%m-%d').date()
                qs = qs.filter(fecha__gte=dfrom)
            except ValueError:
                pass
        if fecha_to:
            try:
                dto = datetime.strptime(fecha_to, '%Y-%m-%d').date()
                qs = qs.filter(fecha__lte=dto)
            except ValueError:
                pass

        count = qs.count()
        if count == 0:
            return Response({'deleted': 0, 'detail': 'No se encontraron registros para eliminar'}, status=status.HTTP_200_OK)

        qs.delete()
        return Response({'deleted': count}, status=status.HTTP_200_OK)


from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class DashboardSalesCorporativoExportProd15(APIView):
    """
    GET: exporta registros filtrados de la empresa del usuario a xlsx.
    Soporta filtros: mes_nombre, nombre_cliente, marca, fecha_from, fecha_to
    """

    def _apply_filters_export_prod15(self, qs, request):
        q = qs
        mes_nombre = request.query_params.get('mes_nombre')
        nombre_cliente = request.query_params.get('nombre_cliente')
        marca = request.query_params.get('marca')
        fecha_from = request.query_params.get('fecha_from')
        fecha_to = request.query_params.get('fecha_to')

        if mes_nombre:
            q = q.filter(mes_nombre__iexact=mes_nombre)
        if nombre_cliente:
            q = q.filter(nombre_cliente__icontains=nombre_cliente)
        if marca:
            q = q.filter(marca__icontains=marca)
        if fecha_from:
            try:
                dfrom = datetime.strptime(fecha_from, '%Y-%m-%d').date()
                q = q.filter(fecha__gte=dfrom)
            except ValueError:
                pass
        if fecha_to:
            try:
                dto = datetime.strptime(fecha_to, '%Y-%m-%d').date()
                q = q.filter(fecha__lte=dto)
            except ValueError:
                pass
        return q

    def get(self, request):
        try:
            usuario = get_usuario_from_token_prod15(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        qs = DashboardSalesCorporativo.objects.filter(id_empresa=empresa)
        qs = self._apply_filters_export_prod15(qs, request)

        serializer = DashboardSalesCorporativoSerializerProd15(qs, many=True)
        data = serializer.data

        wb = Workbook()
        ws = wb.active
        ws.title = "DashboardSalesCorporativo"

        if data and len(data) > 0:
            headers = list(data[0].keys())
        else:
            try:
                headers = list(DashboardSalesCorporativoSerializerProd15.Meta.fields)
            except Exception:
                headers = [
                    'id_registro','id','orden_compra','fecha','mes_nombre','categoria_cliente','nombre_cliente',
                    'categoria_producto','marca','producto','estado_cotizacion','unidades','precio_unitario','observaciones'
                ]

        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)

        for row_idx, row in enumerate(data, start=2):
            for col_idx, h in enumerate(headers, start=1):
                val = row.get(h, None)
                ws.cell(row=row_idx, column=col_idx, value=val)

        for i, _ in enumerate(headers, start=1):
            col = get_column_letter(i)
            max_length = 0
            for cell in ws[col]:
                try:
                    if cell.value:
                        s = str(cell.value)
                        if len(s) > max_length:
                            max_length = len(s)
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[col].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        empresa_id = getattr(empresa, 'id_empresa', None) or getattr(empresa, 'pk', None) or 'empresa'
        today = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"dashboard_salescorporativo_{empresa_id}_{today}.xlsx"

        from django.http import HttpResponse
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
        return response



#Vista para el crud del dashboard de SALES CORPORATIVO del modelo: DashboardSalesCorporativoMetas
# views.py
import jwt
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import AuthenticationFailed
from django.shortcuts import get_object_or_404
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import DashboardSalesCorporativoMetas, Usuario
from .serializers import DashboardSalesCorporativoMetasProduct15Serializer

class DashboardSalesCorporativoMetasProduct15_ListCreate(APIView):
    """
    GET: lista con filtros (ano, mes, categoria_cliente, nombre_cliente, categoria_producto)
    POST: crea una meta; id_producto se forzara a DEFAULT_PRODUCT_ID desde el serializer.
    """

    def _get_usuario_from_token(self, request):
        auth = request.headers.get('Authorization', '')
        token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else (auth.split(' ')[-1] if auth else '')
        if not token:
            raise AuthenticationFailed('No se proporciono token')
        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
            if not id_usuario:
                raise AuthenticationFailed('Token invalido')
        except jwt.ExpiredSignatureError:
            raise AuthenticationFailed('Token expirado')
        except jwt.InvalidTokenError:
            raise AuthenticationFailed('Token invalido')
        try:
            usuario = Usuario.objects.select_related('id_empresa').get(id_usuario=id_usuario)
            return usuario
        except Usuario.DoesNotExist:
            raise AuthenticationFailed('Usuario no encontrado')

    def _apply_filters(self, qs, request):
        q = qs
        ano = request.query_params.get('ano')
        mes = request.query_params.get('mes')
        categoria_cliente = request.query_params.get('categoria_cliente')
        nombre_cliente = request.query_params.get('nombre_cliente')
        categoria_producto = request.query_params.get('categoria_producto')

        if ano:
            try:
                q = q.filter(ano=int(ano))
            except ValueError:
                pass
        if mes:
            q = q.filter(mes__iexact=mes)
        if categoria_cliente:
            q = q.filter(categoria_cliente__icontains=categoria_cliente)
        if nombre_cliente:
            q = q.filter(nombre_cliente__icontains=nombre_cliente)
        if categoria_producto:
            q = q.filter(categoria_producto__icontains=categoria_producto)
        return q

    def get(self, request):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        qs = DashboardSalesCorporativoMetas.objects.filter(id_empresa=empresa)
        qs = self._apply_filters(qs, request)
        serializer = DashboardSalesCorporativoMetasProduct15Serializer(qs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = DashboardSalesCorporativoMetasProduct15Serializer(data=request.data, context={'empresa': empresa})
        if serializer.is_valid():
            obj = serializer.save()
            return Response(DashboardSalesCorporativoMetasProduct15Serializer(obj).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DashboardSalesCorporativoMetasProduct15_Detail(APIView):
    """
    GET/PUT/PATCH/DELETE por pk para metas; valida empresa.
    """

    def _get_usuario_from_token(self, request):
        auth = request.headers.get('Authorization', '')
        token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else (auth.split(' ')[-1] if auth else '')
        if not token:
            raise AuthenticationFailed('No se proporciono token')
        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
            if not id_usuario:
                raise AuthenticationFailed('Token invalido')
        except jwt.ExpiredSignatureError:
            raise AuthenticationFailed('Token expirado')
        except jwt.InvalidTokenError:
            raise AuthenticationFailed('Token invalido')
        try:
            usuario = Usuario.objects.select_related('id_empresa').get(id_usuario=id_usuario)
            return usuario
        except Usuario.DoesNotExist:
            raise AuthenticationFailed('Usuario no encontrado')

    def get(self, request, pk):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)
        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        obj = get_object_or_404(DashboardSalesCorporativoMetas, pk=pk)
        if obj.id_empresa != empresa:
            return Response({'error': 'No autorizado para ver este registro'}, status=status.HTTP_403_FORBIDDEN)
        serializer = DashboardSalesCorporativoMetasProduct15Serializer(obj)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def _update(self, request, pk, partial):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)
        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        obj = get_object_or_404(DashboardSalesCorporativoMetas, pk=pk)
        if obj.id_empresa != empresa:
            return Response({'error': 'No autorizado para modificar este registro'}, status=status.HTTP_403_FORBIDDEN)

        data = dict(request.data)
        data.pop('id_empresa', None)
        data.pop('id_producto', None)
        serializer = DashboardSalesCorporativoMetasProduct15Serializer(obj, data=data, partial=partial, context={'empresa': empresa})
        if serializer.is_valid():
            updated = serializer.save()
            return Response(DashboardSalesCorporativoMetasProduct15Serializer(updated).data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def delete(self, request, pk):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)
        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        obj = get_object_or_404(DashboardSalesCorporativoMetas, pk=pk)
        if obj.id_empresa != empresa:
            return Response({'error': 'No autorizado para eliminar este registro'}, status=status.HTTP_403_FORBIDDEN)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class DashboardSalesCorporativoMetasProduct15_BulkDelete(APIView):
    """
    POST: borra en masa los registros que coinciden con filtros y pertenecen a la empresa del usuario.
    Soporta filtros por body o query params: ano, mes, categoria_cliente, nombre_cliente, categoria_producto
    Responde: {'deleted': <cantidad>}
    """
    def _get_usuario_from_token(self, request):
        auth = request.headers.get('Authorization', '')
        token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else (auth.split(' ')[-1] if auth else '')
        if not token:
            raise AuthenticationFailed('No se proporciono token')
        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
            if not id_usuario:
                raise AuthenticationFailed('Token invalido')
        except jwt.ExpiredSignatureError:
            raise AuthenticationFailed('Token expirado')
        except jwt.InvalidTokenError:
            raise AuthenticationFailed('Token invalido')
        try:
            usuario = Usuario.objects.select_related('id_empresa').get(id_usuario=id_usuario)
            return usuario
        except Usuario.DoesNotExist:
            raise AuthenticationFailed('Usuario no encontrado')

    def post(self, request):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        params = request.data if request.data else request.query_params
        qs = DashboardSalesCorporativoMetas.objects.filter(id_empresa=empresa)

        ano = params.get('ano')
        mes = params.get('mes')
        categoria_cliente = params.get('categoria_cliente')
        nombre_cliente = params.get('nombre_cliente')
        categoria_producto = params.get('categoria_producto')

        if ano:
            try:
                qs = qs.filter(ano=int(ano))
            except ValueError:
                pass
        if mes:
            qs = qs.filter(mes__iexact=mes)
        if categoria_cliente:
            qs = qs.filter(categoria_cliente__icontains=categoria_cliente)
        if nombre_cliente:
            qs = qs.filter(nombre_cliente__icontains=nombre_cliente)
        if categoria_producto:
            qs = qs.filter(categoria_producto__icontains=categoria_producto)

        count = qs.count()
        if count == 0:
            return Response({'deleted': 0, 'detail': 'No se encontraron registros para eliminar'}, status=status.HTTP_200_OK)
        qs.delete()
        return Response({'deleted': count}, status=status.HTTP_200_OK)


class DashboardSalesCorporativoMetasProduct15_Export(APIView):
    """
    GET: Exporta los registros filtrados de la empresa del usuario como .xlsx
    Soporta filtros: ano, mes, categoria_cliente, nombre_cliente, categoria_producto
    """

    def _get_usuario_from_token(self, request):
        auth = request.headers.get('Authorization', '')
        token = auth.split('Bearer ')[-1] if 'Bearer ' in auth else (auth.split(' ')[-1] if auth else '')
        if not token:
            raise AuthenticationFailed('No se proporciono token')
        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            id_usuario = payload.get('id_usuario')
            if not id_usuario:
                raise AuthenticationFailed('Token invalido')
        except jwt.ExpiredSignatureError:
            raise AuthenticationFailed('Token expirado')
        except jwt.InvalidTokenError:
            raise AuthenticationFailed('Token invalido')
        try:
            usuario = Usuario.objects.select_related('id_empresa').get(id_usuario=id_usuario)
            return usuario
        except Usuario.DoesNotExist:
            raise AuthenticationFailed('Usuario no encontrado')

    def _apply_filters(self, qs, request):
        q = qs
        ano = request.query_params.get('ano')
        mes = request.query_params.get('mes')
        categoria_cliente = request.query_params.get('categoria_cliente')
        nombre_cliente = request.query_params.get('nombre_cliente')
        categoria_producto = request.query_params.get('categoria_producto')

        if ano:
            try:
                q = q.filter(ano=int(ano))
            except ValueError:
                pass
        if mes:
            q = q.filter(mes__iexact=mes)
        if categoria_cliente:
            q = q.filter(categoria_cliente__icontains=categoria_cliente)
        if nombre_cliente:
            q = q.filter(nombre_cliente__icontains=nombre_cliente)
        if categoria_producto:
            q = q.filter(categoria_producto__icontains=categoria_producto)
        return q

    def get(self, request):
        try:
            usuario = self._get_usuario_from_token(request)
        except AuthenticationFailed as e:
            return Response({'error': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if empresa is None:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        qs = DashboardSalesCorporativoMetas.objects.filter(id_empresa=empresa)
        qs = self._apply_filters(qs, request)
        serializer = DashboardSalesCorporativoMetasProduct15Serializer(qs, many=True)
        data = serializer.data

        wb = Workbook()
        ws = wb.active
        ws.title = "DashboardSalesCorporativoMetas"

        if data and len(data) > 0:
            headers = list(data[0].keys())
        else:
            headers = list(DashboardSalesCorporativoMetasProduct15Serializer.Meta.fields)

        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)

        for row_idx, row in enumerate(data, start=2):
            for col_idx, h in enumerate(headers, start=1):
                val = row.get(h, None)
                ws.cell(row=row_idx, column=col_idx, value=val)

        for i, _ in enumerate(headers, start=1):
            col = get_column_letter(i)
            max_length = 0
            for cell in ws[col]:
                try:
                    if cell.value:
                        s = str(cell.value)
                        if len(s) > max_length:
                            max_length = len(s)
                except Exception:
                    pass
            ws.column_dimensions[col].width = min(max_length + 2, 60)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        empresa_id = getattr(empresa, 'id_empresa', None) or getattr(empresa, 'pk', None) or 'empresa'
        today = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"dashboard_salescorporativometas_{empresa_id}_{today}.xlsx"

        from django.http import HttpResponse
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
        return response








#Vistas para Dashboard de Isp
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import DashboardIspVentas
from .serializers import DashboardIspVentasSerializer

class DashboardIspVentas_List(APIView):
    """
    API para obtener los registros de Dashboard ISP Ventas
    Permite filtrar por empresa, año, mes, cliente, plan, etc.
    """
    def get(self, request):
        queryset = DashboardIspVentas.objects.all()

        # Filtros opcionales por parámetros GET
        id_empresa = request.GET.get('id_empresa')
        ano = request.GET.get('ano')
        mes = request.GET.get('mes')
        categoria_cliente = request.GET.get('categoria_cliente')
        ciudad = request.GET.get('ciudad')
        segmento = request.GET.get('segmento')
        estado_suscripcion = request.GET.get('estado_suscripcion')

        if id_empresa:
            queryset = queryset.filter(id_empresa=id_empresa)
        if ano:
            queryset = queryset.filter(ano=ano)
        if mes:
            queryset = queryset.filter(mes__iexact=mes)
        if categoria_cliente:
            queryset = queryset.filter(categoria_cliente__iexact=categoria_cliente)
        if ciudad:
            queryset = queryset.filter(ciudad__iexact=ciudad)
        if segmento:
            queryset = queryset.filter(segmento__iexact=segmento)
        if estado_suscripcion:
            queryset = queryset.filter(estado_suscripcion__iexact=estado_suscripcion)

        serializer = DashboardIspVentasSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

































#CHATBOT
# views.py
import re
import json
import jwt
import requests
import time
from datetime import datetime
from decimal import Decimal

from django.conf import settings
from django.db.models import Sum, F, Value, IntegerField, DecimalField, Q
from django.db.models.functions import Coalesce
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import JSONParser
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from rest_framework.authentication import get_authorization_header

# Importa tus modelos (ajusta si Usuario está en otro app)
from .models import DashboardVentas, Usuario

# -----------------------
# Config desde settings (ya en settings.py via decouple)
# -----------------------
OPENAI_API_KEY = getattr(settings, "OPENAI_API_KEY", None)
OPENAI_API_BASE = getattr(settings, "OPENAI_API_BASE", "https://api.openai.com/v1")
DEFAULT_MODEL = getattr(settings, "OPENAI_DEFAULT_MODEL", "gpt-4o-mini")

# -----------------------
# Helpers para parsing
# -----------------------
SPANISH_MONTHS = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12
}

METRIC_KEYWORDS = {
    "cantidad_vendida": ("cantidad_vendida", "int"),
    "cantidad": ("cantidad_vendida", "int"),
    "unidades": ("cantidad_vendida", "int"),
    "dinero_vendido": ("dinero_vendido", "decimal"),
    "dinero": ("dinero_vendido", "decimal"),
    "monto": ("dinero_vendido", "decimal"),
    "ventas": ("dinero_vendido", "decimal"),
    "numero_transacciones": ("numero_transacciones", "int"),
    "transacciones": ("numero_transacciones", "int"),
    "devoluciones": ("devoluciones", "int"),
    "dinero_devoluciones": ("dinero_devoluciones", "decimal"),
    "descuento_total": ("descuento_total", "decimal"),
    "unidades_promocionadas": ("unidades_promocionadas", "int"),
}

def find_metric(text):
    t = text.lower()
    for k in sorted(METRIC_KEYWORDS.keys(), key=lambda x: -len(x)):
        if k in t:
            return METRIC_KEYWORDS[k]
    return None

def extract_year(text):
    m = re.search(r'\b(19|20)\d{2}\b', text)
    if m:
        try:
            return int(m.group(0))
        except:
            return None
    return None

def extract_month(text):
    t = text.lower()
    for name, num in SPANISH_MONTHS.items():
        if name in t:
            return num
    m = re.search(r'\bmes(?:\s+de)?\s+(\d{1,2})\b', t)
    if m:
        try:
            val = int(m.group(1))
            if 1 <= val <= 12:
                return val
        except:
            pass
    return None

def extract_brand(text):
    m = re.search(r'marca\s+[:\-]?\s*["\']?([A-Za-z0-9\-\_\. ]{1,60})["\']?', text, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return None

def extract_sku(text):
    m = re.search(r'\bsku\s*[:\-]?\s*([A-Za-z0-9\-\_]{1,60})\b', text, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return None

def extract_producto(text):
    m = re.search(r'producto\s+[:\-]?\s*["\']?([A-Za-z0-9\-\_\. ]{1,80})["\']?', text, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return None

# -----------------------
# Vista principal
# -----------------------
class ChatbotAPIView(APIView):
    """
    Endpoint que:
     - valida JWT (usa id_usuario del payload para obtener Usuario)
     - si detecta consulta de métrica + filtros, consulta DashboardVentas filtrando por id_empresa
       del usuario y devuelve SOLO el total agregado en 'reply'.
     - si no detecta métrica, reenvía el mensaje a OpenAI (chat/completions) como fallback.
    """

    parser_classes = [JSONParser]

    @method_decorator(csrf_exempt)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request):
        # Validar token JWT
        auth_header = get_authorization_header(request).split()
        if not auth_header or auth_header[0].lower() != b'bearer':
            return Response({'error': 'Token no enviado'}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            token = auth_header[1].decode('utf-8')
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
        except jwt.ExpiredSignatureError:
            return Response({'error': 'Token expirado'}, status=status.HTTP_401_UNAUTHORIZED)
        except jwt.DecodeError:
            return Response({'error': 'Token inválido'}, status=status.HTTP_401_UNAUTHORIZED)
        except Exception as e:
            return Response({'error': 'Error validando token', 'detail': str(e)}, status=status.HTTP_401_UNAUTHORIZED)

        id_usuario = payload.get('id_usuario')
        if not id_usuario:
            return Response({'error': 'Token no contiene id_usuario'}, status=status.HTTP_401_UNAUTHORIZED)

        # Obtener usuario y su empresa
        try:
            usuario = Usuario.objects.select_related('id_empresa').get(id_usuario=id_usuario)
        except Usuario.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_401_UNAUTHORIZED)
        except Exception as e:
            return Response({'error': 'Error buscando usuario', 'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        empresa = getattr(usuario, 'id_empresa', None)
        if not empresa:
            return Response({'error': 'Usuario no pertenece a ninguna empresa'}, status=status.HTTP_403_FORBIDDEN)

        # Body
        data = request.data
        user_message = (data.get('message') or "").strip()
        if not user_message:
            return Response({'error': 'No se envió message'}, status=status.HTTP_400_BAD_REQUEST)

        # Intent: detectar métrica y filtros
        metric = find_metric(user_message)  # (campo, tipo) o None
        year = extract_year(user_message)
        month = extract_month(user_message)
        brand = extract_brand(user_message)
        sku = extract_sku(user_message)
        producto = extract_producto(user_message)

        # Si detectamos métrica -> hacemos consulta a la DB
        if metric:
            campo, tipo = metric
            # Construir filtros
            q = Q(id_empresa=empresa)

            if year:
                q &= Q(anio=year)
            if month:
                q &= Q(mes=month)
            if brand:
                q &= Q(marca__icontains=brand)
            if sku:
                q &= Q(sku__iexact=sku)
            if producto:
                q &= Q(nombre_producto__icontains=producto)

            # Agregación
            try:
                if tipo == 'int':
                    output_field = IntegerField()
                    zero_value = Value(0, output_field=IntegerField())
                    agg = DashboardVentas.objects.filter(q).aggregate(
                        total=Coalesce(
                            Sum(F(campo), output_field=output_field),
                            zero_value,
                            output_field=output_field
                        )
                    )
                    total = agg.get('total')
                    try:
                        total_int = int(total) if total is not None else 0
                        reply_value = str(total_int)
                    except Exception:
                        reply_value = str(total)
                else:
                    output_field = DecimalField(max_digits=20, decimal_places=2)
                    zero_value = Value(Decimal('0.00'), output_field=output_field)
                    agg = DashboardVentas.objects.filter(q).aggregate(
                        total=Coalesce(
                            Sum(F(campo), output_field=output_field),
                            zero_value,
                            output_field=output_field
                        )
                    )
                    total = agg.get('total')
                    if total is None:
                        total = Decimal('0.00')
                    try:
                        reply_value = f"{Decimal(total):.2f}"
                    except Exception:
                        reply_value = str(total)
            except Exception as e:
                return Response({'error': 'Error al agregar datos', 'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            # Responder SOLO con el total en `reply`
            return Response({
                "reply": reply_value,
                "metric": campo,
                "filters": {
                    "anio": year,
                    "mes": month,
                    "marca": brand,
                    "sku": sku,
                    "producto": producto
                },
                "total": reply_value
            }, status=status.HTTP_200_OK)

        # -----------------------
        # Fallback: reenviar a OpenAI (si no detectamos métrica)
        # -----------------------
        if not OPENAI_API_KEY:
            return Response({'error': 'Falta configuración de OPENAI_API_KEY en settings/.env'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        payload = {
            "model": data.get("model", DEFAULT_MODEL),
            "messages": [
                {"role": "user", "content": user_message}
            ],
            "max_tokens": data.get("max_tokens", 800),
            "temperature": data.get("temperature", 0.2),
        }

        headers = {
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json",
        }

        try:
            resp = requests.post(
                f"{OPENAI_API_BASE}/chat/completions",
                headers=headers,
                json=payload,
                timeout=30
            )
        except requests.RequestException as e:
            return Response({'error': 'Error conectando con la API externa', 'detail': str(e)}, status=status.HTTP_502_BAD_GATEWAY)

        if resp.status_code != 200:
            try:
                err_body = resp.json()
            except Exception:
                err_body = resp.text
            return Response({'error': 'Error desde la API de OpenAI', 'status': resp.status_code, 'detail': err_body}, status=status.HTTP_502_BAD_GATEWAY)

        try:
            body = resp.json()
            choices = body.get("choices", [])
            if not choices:
                return Response({'error': 'No choices devueltos por OpenAI', 'raw': body}, status=status.HTTP_502_BAD_GATEWAY)

            assistant_text = choices[0].get("message", {}).get("content", "")
            return Response({
                "reply": assistant_text,
                "raw": body
            }, status=status.HTTP_200_OK)
        except ValueError:
            return Response({'error': 'Respuesta inválida de OpenAI', 'raw': resp.text}, status=status.HTTP_502_BAD_GATEWAY)










 #DASHBOARD CHURN RATE PARA SERVITEL
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.utils.dateparse import parse_date

from .models import DashboardChurnRate
from .serializers import DashboardChurnRateSerializer

class DashboardChurnRateView(APIView):
    """
    Listado de DashboardChurnRate filtrado automáticamente por id_empresa del usuario autenticado.
    Permisos: solo usuarios autenticados (IsAuthenticated).
    Query params opcionales:
      - start : fecha mínima para fecha_ultima_transaccion (YYYY-MM-DD)
      - end   : fecha máxima para fecha_ultima_transaccion (YYYY-MM-DD)
      - estado: filtrar por estado_cliente (ej. activo, cancelado, inactivo)
      - tipo   : filtrar por tipo_plan (ej. basico, estandar, premium)
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        usuario = request.user

        # Verificamos que el user tenga id_empresa (sigue tu convención de usuario)
        if not hasattr(usuario, 'id_empresa') or usuario.id_empresa is None:
            return Response({'error': 'Usuario inválido en request o sin id_empresa.'},
                            status=status.HTTP_401_UNAUTHORIZED)

        # Filtrado base por la empresa del usuario
        queryset = DashboardChurnRate.objects.filter(id_empresa=usuario.id_empresa).order_by('-fecha_ultima_transaccion')

        # Filtros opcionales por query params
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        if start:
            fecha = parse_date(start)
            if fecha:
                queryset = queryset.filter(fecha_ultima_transaccion__gte=fecha)
        if end:
            fecha = parse_date(end)
            if fecha:
                queryset = queryset.filter(fecha_ultima_transaccion__lte=fecha)

        estado = request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado_cliente=estado)

        tipo = request.query_params.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo_plan=tipo)

        serializer = DashboardChurnRateSerializer(queryset, many=True, context={'usuario': usuario})
        return Response(serializer.data, status=status.HTTP_200_OK)








#----------------------Dashboard ISP ARPU-------------------------------

# views.py
# views.py (version sin autenticacion)
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.utils.dateparse import parse_date
from django.db import transaction

from .models import DashboardARPU
from .serializers import DashboardARPUSerializer

from datetime import datetime
import numpy as np
import pandas as pd
from statsmodels.tsa.holtwinters import ExponentialSmoothing

class DashboardARPUListView(APIView):
    """
    GET: lista registros DashboardARPU filtrados por parametros.
    Parametros opcionales: start (YYYY-MM-DD), end (YYYY-MM-DD), producto (id), empresa (id)
    """

    def get(self, request):
        qs = DashboardARPU.objects.all().order_by('periodo_mes')

        start = request.query_params.get('start')
        end = request.query_params.get('end')
        producto = request.query_params.get('producto')
        empresa = request.query_params.get('empresa')

        if empresa:
            qs = qs.filter(id_empresa_id=empresa)
        if producto:
            qs = qs.filter(id_producto_id=producto)
        if start:
            try:
                qs = qs.filter(periodo_mes__gte=parse_date(start))
            except Exception:
                return Response({'error': 'start no es una fecha valida (YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)
        if end:
            try:
                qs = qs.filter(periodo_mes__lte=parse_date(end))
            except Exception:
                return Response({'error': 'end no es una fecha valida (YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = DashboardARPUSerializer(qs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class DashboardARPUForecastView(APIView):
    """
    GET: calculo de forecast on-the-fly para una combinacion (empresa, producto).
    Params requeridos: producto, empresa. Opcional: periods (int) default 6.
    """

    def get(self, request):
        empresa = request.query_params.get('empresa')
        producto = request.query_params.get('producto')
        if not empresa or not producto:
            return Response({'error': 'parametros empresa y producto son requeridos'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            empresa_id = int(empresa)
            producto_id = int(producto)
        except Exception:
            return Response({'error': 'empresa y producto deben ser enteros'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            periods = int(request.query_params.get('periods', 6))
        except Exception:
            periods = 6

        qs = DashboardARPU.objects.filter(id_empresa_id=empresa_id, id_producto_id=producto_id).order_by('periodo_mes')
        if not qs.exists():
            return Response({'error': 'No hay datos historicos para esa combinacion empresa/producto'}, status=status.HTTP_404_NOT_FOUND)

        df = pd.DataFrame.from_records(qs.values('periodo_mes', 'arpu'))
        df['periodo_mes'] = pd.to_datetime(df['periodo_mes'])
        df = df.set_index('periodo_mes').sort_index()

        if df['arpu'].isnull().all():
            return Response({'error': 'serie arpu vacia'}, status=status.HTTP_400_BAD_REQUEST)
        df['arpu'] = df['arpu'].fillna(method='ffill').fillna(method='bfill').fillna(0)
        series = df['arpu'].astype(float)

        try:
            seasonal = 'add' if len(series) >= 24 else None
            seasonal_periods = 12 if seasonal else None
            model = ExponentialSmoothing(
                series,
                trend='add',
                seasonal=seasonal,
                seasonal_periods=seasonal_periods,
                initialization_method='estimated'
            )
            fit = model.fit(optimized=True)
            forecast_vals = fit.forecast(steps=periods)
            residuals = fit.resid
            sigma = residuals.std(ddof=1) if len(residuals) > 1 else max(1.0, series.std() if len(series) > 1 else 1.0)
        except Exception as e:
            try:
                model = ExponentialSmoothing(series, trend='add', seasonal=None, initialization_method='estimated')
                fit = model.fit(optimized=True)
                forecast_vals = fit.forecast(steps=periods)
                residuals = fit.resid
                sigma = residuals.std(ddof=1) if len(residuals) > 1 else max(1.0, series.std() if len(series) > 1 else 1.0)
            except Exception as e2:
                return Response({'error': 'Error al entrenar modelo statsmodels: %s' % str(e2)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        last_index = series.index.max()
        preds = []
        for i, val in enumerate(forecast_vals, start=1):
            periodo = (last_index + pd.DateOffset(months=i)).to_pydatetime().date().replace(day=1)
            lower = float(val) - 1.96 * float(sigma)
            upper = float(val) + 1.96 * float(sigma)
            preds.append({
                'periodo': periodo.isoformat(),
                'arpu_pred': round(float(val), 2),
                'lower': round(max(0.0, lower), 2),
                'upper': round(max(0.0, upper), 2),
            })

        response = {
            'empresa': empresa_id,
            'producto': producto_id,
            'horizonte': periods,
            'modelo': 'ExponentialSmoothing_hw' if seasonal else 'ExponentialSmoothing_trend',
            'trained_at': datetime.utcnow().isoformat() + 'Z',
            'predicciones': preds
        }
        return Response(response, status=status.HTTP_200_OK)

class DashboardARPUUpsertView(APIView):
    """
    POST: crea o actualiza (upsert) un registro DashboardARPU por (id_empresa, id_producto, periodo_mes).
    """

    @transaction.atomic
    def post(self, request):
        data = request.data.copy()

        periodo = data.get('periodo_mes')
        if not periodo:
            return Response({'error': 'periodo_mes es requerido (YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            periodo_date = parse_date(periodo) if isinstance(periodo, str) else periodo
            if periodo_date is None:
                raise ValueError()
        except Exception:
            return Response({'error': 'periodo_mes no es una fecha valida (YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)

        producto = data.get('id_producto') or data.get('id_producto_id')
        empresa = data.get('id_empresa') or data.get('id_empresa_id')
        if not producto or not empresa:
            return Response({'error': 'id_producto e id_empresa son requeridos'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            producto_id = int(producto)
            empresa_id = int(empresa)
        except Exception:
            return Response({'error': 'id_producto e id_empresa deben ser enteros'}, status=status.HTTP_400_BAD_REQUEST)

        instance = DashboardARPU.objects.filter(id_empresa_id=empresa_id, id_producto_id=producto_id, periodo_mes=periodo_date).first()

        if instance:
            serializer = DashboardARPUSerializer(instance, data=data, partial=True)
        else:
            serializer = DashboardARPUSerializer(data=data)

        if not serializer.is_valid():
            return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        obj = serializer.save()
        return Response(DashboardARPUSerializer(obj).data, status=status.HTTP_201_CREATED)



























#Formulario de creación
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticatedOrReadOnly, AllowAny, IsAuthenticated
from django.shortcuts import get_object_or_404
from django.db import transaction
from .models import Formulario, Pregunta, Respuesta
from .serializers import FormularioCreateSerializer, FormularioDetailSerializer, RespuestaSerializer

class FormularioViewSet(viewsets.ModelViewSet):
    """
    ViewSet para formularios:
    - create/list/retrieve/update/destroy (lookup por slug)
    - submit (public) -> /api/formularios/{slug}/submit/
    - respuestas (auth'd) -> /api/formularios/{slug}/respuestas/
    """
    queryset = Formulario.objects.all()
    lookup_field = 'slug'
    permission_classes = [IsAuthenticatedOrReadOnly]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return FormularioCreateSerializer
        return FormularioDetailSerializer

    def perform_create(self, serializer):
        """
        Asigna automaticamente empresa y usuario desde request.user si no vienen en payload.
        request.user debe ser tu modelo Usuario (el que tiene id_empresa).
        """
        request = self.request
        user = getattr(request, 'user', None)

        empresa_obj = None
        if user:
            empresa_obj = getattr(user, 'id_empresa', None)

        usuario_obj = None
        if user:
            usuario_obj = user if hasattr(user, 'id_usuario') else None

        if empresa_obj and usuario_obj:
            serializer.save(empresa=empresa_obj, usuario=usuario_obj)
        elif empresa_obj:
            serializer.save(empresa=empresa_obj)
        elif usuario_obj:
            serializer.save(usuario=usuario_obj)
        else:
            serializer.save()

    def _evaluate_branching_rule(self, pregunta, answer_value):
        """
        Dada una pregunta (instancia Pregunta) y el valor respondido, devolvemos:
           - 'end'  => terminar
           - index (int) => índice destino
           - None => sin regla aplicable
        Asume pregunta.branching es un array de { when, goto }.
        Para checkbox (answer_value: array) devolvemos la primera regla que coincida con alguna opción marcada.
        """
        rules = getattr(pregunta, 'branching', None) or []
        if not rules:
            return None

        # checkbox: comprobar si alguna opción marcada coincide con 'when' de una regla
        if pregunta.tipo == 'checkbox' and isinstance(answer_value, (list, tuple)):
            for rule in rules:
                if rule is None:
                    continue
                when = rule.get('when', None)
                goto = rule.get('goto', None)
                if when is None or when == '':
                    continue
                # si any option equals when
                for marked in answer_value:
                    if str(marked) == str(when):
                        return goto
            return None

        # para resto de tipos: comparar igualdad como strings
        for rule in rules:
            if rule is None:
                continue
            when = rule.get('when', None)
            goto = rule.get('goto', None)
            if when is None or when == '':
                continue
            if str(when) == str(answer_value):
                return goto
        return None

    @action(detail=True, methods=['post'], permission_classes=[AllowAny], url_path='submit')
    def submit(self, request, slug=None):
        """
        Endpoint público para recibir respuestas.

        Cambios clave:
        - Simula el flujo empezando en la primera pregunta (orden asc).
        - Aplica branching usando las respuestas recibidas (payload.data).
        - Valida sólo las preguntas que fueron "visitadas" en el flujo.
        - Guarda en Respuesta.data únicamente las respuestas de las preguntas visitadas (con claves legibles).
        """
        formulario = get_object_or_404(Formulario, slug=slug)
        payload = request.data or {}
        answers = payload.get('data', {}) or {}
        errors = {}
        stored = {}

        # Ordenar preguntas por 'orden' y construir lista indexable por posición
        preguntas_qs = list(formulario.preguntas.all().order_by('orden'))
        if not preguntas_qs:
            return Response({'error': 'Formulario sin preguntas'}, status=status.HTTP_400_BAD_REQUEST)

        # Mapa id_pregunta -> pregunta (instancia)
        preguntas_by_id = {str(p.id_pregunta): p for p in preguntas_qs}

        # Obtener id_empresa desde request.user o formulario
        id_empresa_value = None
        user = getattr(request, 'user', None)
        if user:
            id_empresa_value = getattr(user, 'id_empresa_id', None) or getattr(user, 'id_empresa', None)
            if hasattr(id_empresa_value, 'id_empresa'):
                id_empresa_value = id_empresa_value.id_empresa
        if id_empresa_value is None:
            id_empresa_value = getattr(formulario, 'empresa_id', None)

        # Simulamos el recorrido desde el primer índice (0) hasta terminar
        visited_question_ids = []  # recoger id_pregunta (int) que fueron mostradas
        visited_indices = set()
        current_index = 0
        # para evitar loops infinitos: max iter = len(preguntas_qs) * 5 (arbitrario) o detectar revisitas
        max_steps = len(preguntas_qs) * 5
        step = 0
        while True:
            step += 1
            if step > max_steps:
                # protección contra ciclos
                return Response({'error': 'Error en reglas de ramificación (posible bucle infinito).'}, status=status.HTTP_400_BAD_REQUEST)

            if current_index is None or current_index < 0 or current_index >= len(preguntas_qs):
                break

            pregunta = preguntas_qs[current_index]
            pid_str = str(pregunta.id_pregunta)

            # marcar visitada
            visited_question_ids.append(pregunta.id_pregunta)
            visited_indices.add(current_index)

            # obtener respuesta (si existe)
            answer_present = pid_str in answers and answers.get(pid_str) not in [None, "", []]
            answer_value = answers.get(pid_str) if answer_present else None

            # Validar requeridos SOLO si la pregunta fue mostrada (estamos en el flujo) y es requerida
            if pregunta.requerido and not answer_present:
                errors[pid_str] = 'Campo requerido'
                # no rompemos aún: se devolverá al final

            # Evaluar branching usando la respuesta (si no hay respuesta, no habrá coincidencia y se seguirá secuencialmente)
            goto = None
            if answer_present:
                goto = self._evaluate_branching_rule(pregunta, answer_value)

            if goto == 'end':
                # terminamos flujo
                break

            if goto is not None and goto != '':
                # goto debería indicar un índice (0-based) según diseño frontend
                try:
                    target_index = int(goto)
                    if target_index < 0 or target_index >= len(preguntas_qs):
                        # índice inválido, terminamos el flujo por seguridad
                        break
                    # prevenir loops sencillos: si ya visitamos target muchas veces, abortamos (detección simple)
                    if target_index in visited_indices and step > len(preguntas_qs) * 2:
                        return Response({'error': 'Bucle detectado en ramificación.'}, status=status.HTTP_400_BAD_REQUEST)
                    current_index = target_index
                    continue
                except Exception:
                    # si goto no es convertible a int, ignoramos y seguimos secuencialmente
                    current_index = current_index + 1
                    if current_index >= len(preguntas_qs):
                        break
                    continue

            # si no hay regla aplicable, vamos secuencialmente a la siguiente pregunta
            current_index = current_index + 1
            if current_index >= len(preguntas_qs):
                break

        # Si hubo errores de validación en preguntas visitadas, responder con esos errores
        # Convertimos claves a id_pregunta (string) en el dict errors (ya lo guardamos así)
        if errors:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        # Construimos stored sólo con las preguntas visitadas y sus respuestas (si las tienen)
        for pregunta in preguntas_qs:
            if pregunta.id_pregunta not in visited_question_ids:
                continue
            pid_str = str(pregunta.id_pregunta)
            if pid_str not in answers:
                continue
            raw = answers.get(pid_str)
            # normalizamos por tipo (mismo comportamiento previo)
            if pregunta.tipo == 'int':
                try:
                    val = int(raw)
                except Exception:
                    # si invalido pero no fue requerido (o validado antes), dejar como raw
                    try:
                        val = int(float(raw))
                    except Exception:
                        val = raw
            elif pregunta.tipo == 'float':
                try:
                    val = float(raw)
                except Exception:
                    val = raw
            elif pregunta.tipo == 'checkbox':
                val = raw if isinstance(raw, (list, tuple)) else [raw]
            else:
                val = raw

            question_key = pregunta.texto.strip() if pregunta.texto else f"pregunta_{pid_str}"
            stored[question_key] = val

        # Agregar id_empresa
        if id_empresa_value is not None:
            stored_with_empresa = {'id_empresa': id_empresa_value, **stored}
        else:
            stored_with_empresa = stored

        # Guardar respuesta
        with transaction.atomic():
            respuesta = Respuesta.objects.create(formulario=formulario, data=stored_with_empresa)
            serializer = RespuestaSerializer(respuesta)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'], permission_classes=[IsAuthenticated], url_path='respuestas')
    def listar_respuestas(self, request, slug=None):
        usuario = request.user
        formulario = get_object_or_404(Formulario, slug=slug)

        user_empresa_id = getattr(usuario, 'id_empresa_id', None)
        if user_empresa_id is None:
            try:
                user_empresa_id = usuario.id_empresa.id_empresa
            except Exception:
                user_empresa_id = None

        if user_empresa_id is None or formulario.empresa_id != user_empresa_id:
            return Response({'error': 'No autorizado para ver respuestas'}, status=status.HTTP_403_FORBIDDEN)

        qs = formulario.respuestas.all().order_by('-fecha')
        serializer = RespuestaSerializer(qs, many=True)
        return Response(serializer.data)








#CHATBOT DE N8N# myapp/serializers.py
# myapp/views.py
import time
import jwt
import requests
from requests.exceptions import RequestException, Timeout
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from django.conf import settings

from .serializers import WebhookProxySerializer, DashboardContextSerializer
from .models import DashboardContext, Usuario

# --- Configuración (lee de settings.py) ---
TARGET_WEBHOOK_URL = getattr(settings, "CHAT_TARGET_WEBHOOK_URL", None)
WEBHOOK_JWT_SECRET = getattr(settings, "WEBHOOK_JWT_SECRET", None)
WEBHOOK_JWT_ALGORITHM = getattr(settings, "WEBHOOK_JWT_ALGORITHM", "HS256")
WEBHOOK_JWT_EXP_SECONDS = int(getattr(settings, "WEBHOOK_JWT_EXP_SECONDS", 3600))

N8N_REQUEST_TIMEOUT = float(getattr(settings, "N8N_REQUEST_TIMEOUT", 120.0))
N8N_MAX_RETRIES = int(getattr(settings, "N8N_MAX_RETRIES", 1))
N8N_BACKOFF_FACTOR = float(getattr(settings, "N8N_BACKOFF_FACTOR", 0.5))

LOGIN_TOKEN_ALGORITHM = "HS256"  # algoritmo del token de login (el que llega en Authorization)

# --- Helpers ---
def _make_jwt_for_webhook():
    now = int(time.time())
    payload = {"sub": "webhook", "iat": now, "exp": now + WEBHOOK_JWT_EXP_SECONDS}
    token = jwt.encode(payload, WEBHOOK_JWT_SECRET, algorithm=WEBHOOK_JWT_ALGORITHM)
    return token.decode("utf-8") if isinstance(token, bytes) else token

def _requests_session_with_retries(retries=N8N_MAX_RETRIES, backoff=N8N_BACKOFF_FACTOR):
    session = requests.Session()
    retry = Retry(
        total=retries,
        read=retries,
        connect=retries,
        backoff_factor=backoff,
        status_forcelist=(429, 502, 503, 504),
        allowed_methods=frozenset(["POST", "GET"])
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session

def _extract_token_from_header(request):
    auth_header = request.headers.get("Authorization", "")
    if not auth_header or not auth_header.startswith("Bearer "):
        return None
    return auth_header.split(" ", 1)[1].strip()

def _validate_and_get_usuario(token):
    try:
        payload = jwt.decode(token, settings.SECRET_KEY, algorithms=[LOGIN_TOKEN_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise ValueError("Token expirado")
    except jwt.InvalidTokenError:
        raise ValueError("Token inválido")

    id_usuario = payload.get("id_usuario")
    if not id_usuario:
        raise ValueError("Token sin id_usuario")

    try:
        usuario = Usuario.objects.select_related("id_empresa", "id_estado").get(id_usuario=id_usuario)
    except Usuario.DoesNotExist:
        raise ValueError("Usuario no encontrado")

    try:
        estado_id = usuario.id_estado.id_estado
    except Exception:
        estado_id = None
    if estado_id != 1:
        raise PermissionError("Usuario inactivo")

    return usuario

# --- API: listar DashboardContext para la empresa del usuario ---
class DashboardContextListAPIView(APIView):
    """
    GET /n8n/dashboard-contexts/
    - Requiere Authorization: Bearer <token_de_login>
    - Devuelve la lista de DashboardContext asociados a la empresa del usuario.
    """
    def get(self, request, *args, **kwargs):
        try:
            token = _extract_token_from_header(request)
            if not token:
                return Response({"detail": "Authorization header missing or malformed"}, status=status.HTTP_401_UNAUTHORIZED)

            try:
                usuario = _validate_and_get_usuario(token)
            except ValueError as ve:
                return Response({"detail": str(ve)}, status=status.HTTP_401_UNAUTHORIZED)
            except PermissionError as pe:
                return Response({"detail": str(pe)}, status=status.HTTP_403_FORBIDDEN)

            empresa_id_val = usuario.id_empresa.id_empresa
            queryset = DashboardContext.objects.filter(empresa_id=empresa_id_val).order_by("id_registro")
            serializer = DashboardContextSerializer(queryset, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"detail": "Error interno", "error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# --- API: proxy a n8n (recibe id_registro + chatInput) ---
class ChatWebhookProxyAPIView(APIView):
    """
    POST /n8n/webhook-proxy/
    Body: { id_registro: int, chatInput: str }
    - Valida token de login (Authorization Bearer ...)
    - Busca DashboardContext por id_registro y empresa del usuario
    - Forma payload usando LOS CAMPOS DE LA TABLA (session_id, dashboard_name, dashboard_context, tables, formularios_id, empresa_id)
      y usa chatInput enviado por el cliente (no utiliza chat_input almacenado en la BD).
    - Firma con JWT propio y reenvía a TARGET_WEBHOOK_URL.
    """
    def post(self, request, *args, **kwargs):
        try:
            if not TARGET_WEBHOOK_URL or not WEBHOOK_JWT_SECRET:
                return Response({"detail": "Configuración incompleta del webhook"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            token = _extract_token_from_header(request)
            if not token:
                return Response({"detail": "Authorization header missing or malformed"}, status=status.HTTP_401_UNAUTHORIZED)

            try:
                usuario = _validate_and_get_usuario(token)
            except ValueError as ve:
                return Response({"detail": str(ve)}, status=status.HTTP_401_UNAUTHORIZED)
            except PermissionError as pe:
                return Response({"detail": str(pe)}, status=status.HTTP_403_FORBIDDEN)

            # Validar body
            serializer = WebhookProxySerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            validated = serializer.validated_data

            id_registro = validated["id_registro"]
            chat_input = validated["chatInput"]

            # Obtener contexto (asegurando que pertenezca a la misma empresa)
            try:
                context = DashboardContext.objects.get(id_registro=id_registro, empresa_id=usuario.id_empresa.id_empresa)
            except DashboardContext.DoesNotExist:
                return Response({"detail": "DashboardContext no encontrado para esta empresa"}, status=status.HTTP_404_NOT_FOUND)

            # Formar payload para n8n (usa campos de la tabla, pero chatInput viene del cliente)
            payload_to_send = {
                "chatInput": chat_input,
                "sessionId": context.session_id,
                "dashboard_name": context.dashboard_name,
                "dashboard_context": context.dashboard_context,
                "tables": context.tables,
                "formularios_id": context.formularios_id or {},
                "empresaId": context.empresa_id
            }

            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {_make_jwt_for_webhook()}"
            }

            session = _requests_session_with_retries()

            try:
                response = session.post(
                    TARGET_WEBHOOK_URL,
                    json=payload_to_send,
                    headers=headers,
                    timeout=N8N_REQUEST_TIMEOUT
                )
            except Timeout:
                return Response({"detail": "Timeout al contactar n8n"}, status=status.HTTP_504_GATEWAY_TIMEOUT)
            except RequestException as e:
                return Response({"detail": "Error contactando n8n", "error": str(e)}, status=status.HTTP_502_BAD_GATEWAY)

            try:
                data = response.json()
            except ValueError:
                data = response.text

            return Response({"response_from_webhook": data}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"detail": "Error interno", "error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)













# LISTADO DE LOS FORMUYLARIOS
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from .models import Formulario
from .serializers import ListadoFormulariosSerializer

class ListadoFormulariosView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        usuario = request.user

        if not hasattr(usuario, 'id_usuario'):
            return Response({'error': 'Usuario inválido'}, status=status.HTTP_401_UNAUTHORIZED)

        empresa = getattr(usuario, 'id_empresa', None)
        if not empresa:
            return Response({'error': 'Usuario sin empresa asignada'}, status=status.HTTP_400_BAD_REQUEST)

        formularios = Formulario.objects.filter(
            empresa=empresa
        ).order_by('-fecha_creacion')

        serializer = ListadoFormulariosSerializer(formularios, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)




#Editar formulario

# your_app/views.py
from django.shortcuts import get_object_or_404
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from .models import Formulario, Pregunta
from .serializers import FormularioEditSerializer

class FormularioEditView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request, slug):
        usuario = request.user
        if not hasattr(usuario, 'id_usuario'):
            return Response({'error': 'Usuario inválido'}, status=status.HTTP_401_UNAUTHORIZED)

        formulario = get_object_or_404(Formulario, slug=slug)
        empresa_usuario = getattr(usuario, 'id_empresa', None)
        if empresa_usuario is None or formulario.empresa.id_empresa != empresa_usuario.id_empresa:
            return Response({'error': 'No autorizado para este formulario'}, status=status.HTTP_403_FORBIDDEN)

        serializer = FormularioEditSerializer(formulario, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, slug):
        """
        Se espera payload:
        {
          "nombre": "Nuevo nombre",
          "descripcion": "nueva desc",
          "preguntas": [
             {"texto":"P1", "tipo":"text", "orden":0, "requerido":false, "opciones":[], "branching":[]},
             ...
          ]
        }
        La operación reemplaza las preguntas actuales por las nuevas.
        """
        usuario = request.user
        if not hasattr(usuario, 'id_usuario'):
            return Response({'error': 'Usuario inválido'}, status=status.HTTP_401_UNAUTHORIZED)

        formulario = get_object_or_404(Formulario, slug=slug)
        empresa_usuario = getattr(usuario, 'id_empresa', None)
        if empresa_usuario is None or formulario.empresa.id_empresa != empresa_usuario.id_empresa:
            return Response({'error': 'No autorizado para este formulario'}, status=status.HTTP_403_FORBIDDEN)

        data = request.data
        nombre = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', None)
        preguntas_payload = data.get('preguntas', [])

        if not nombre:
            return Response({'error': 'El nombre es requerido'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                # actualizar formulario
                formulario.nombre = nombre
                formulario.descripcion = descripcion
                formulario.save()

                # eliminar preguntas previas y crear nuevas (reemplazo completo)
                Pregunta.objects.filter(formulario=formulario).delete()

                preguntas_creadas = []
                for idx, p in enumerate(preguntas_payload):
                    texto = p.get('texto', '').strip() or f'Pregunta {idx+1}'
                    tipo = p.get('tipo', 'text')
                    orden = p.get('orden', idx)
                    requerido = bool(p.get('requerido', False))
                    opciones = p.get('opciones', None)
                    branching = p.get('branching', None)

                    pregunta = Pregunta.objects.create(
                        formulario=formulario,
                        texto=texto,
                        tipo=tipo,
                        orden=orden,
                        requerido=requerido,
                        opciones=opciones if opciones not in ([], None) else None,
                        branching=branching if branching not in ([], None) else None,
                    )
                    preguntas_creadas.append(pregunta)

                # respuesta con estado actualizado
                serializer = FormularioEditSerializer(formulario, context={'request': request})
                return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            # loguea si tienes logger configurado (no incluido aquí para mantenerlo simple)
            return Response({'error': 'Error al actualizar formulario', 'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)








#Serializador para dashboard de ventas de formulario de ventas espacio y mercadeo





# backend/appdataflowai/views.py
# backend/appdataflowai/views.py
from django.utils.dateparse import parse_date
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from .models import Respuesta
from .serializers import DashboardFormsVentasPuntoVentaSAerializer
import logging

logger = logging.getLogger(__name__)

class DashboardFormsVentasPuntoVentaView(APIView):
    """
    Devuelve respuestas del formulario (por defecto id_formulario = 15).
    Requiere token (IsAuthenticated).
    Query params:
      - form_id (opcional) -> id_formulario (por defecto 15)
      - start (opcional) -> fecha desde (YYYY-MM-DD)
      - end (opcional) -> fecha hasta (YYYY-MM-DD)
    """
    permission_classes = (IsAuthenticated,)

    DEFAULT_FORM_ID = 15

    def get(self, request):
        usuario = request.user
        if not hasattr(usuario, 'id_usuario'):
            return Response({'error': 'Usuario inválido en request'}, status=status.HTTP_401_UNAUTHORIZED)

        # Obtener form_id desde query params o usar default
        form_id_param = request.query_params.get('form_id')
        try:
            form_id = int(form_id_param) if form_id_param is not None else self.DEFAULT_FORM_ID
        except (ValueError, TypeError):
            return Response({'error': 'form_id inválido. Debe ser entero.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            empresa_fk = getattr(usuario, 'id_empresa', None)
            queryset = Respuesta.objects.filter(formulario__id_formulario=form_id)

            if empresa_fk is not None:
                queryset = queryset.filter(formulario__empresa=empresa_fk)

            # Filtros de fecha (start / end)
            start = request.query_params.get('start')
            end = request.query_params.get('end')

            if start:
                date_start = parse_date(start)
                if not date_start:
                    return Response({'error': 'Formato de fecha "start" inválido. Use YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)
                queryset = queryset.filter(fecha__date__gte=date_start)

            if end:
                date_end = parse_date(end)
                if not date_end:
                    return Response({'error': 'Formato de fecha "end" inválido. Use YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)
                queryset = queryset.filter(fecha__date__lte=date_end)

            queryset = queryset.order_by('-fecha')

        except Exception as exc:
            logger.exception("Error al consultar respuestas del formulario %s", form_id)
            return Response({'error': 'Error al consultar datos'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        serializer = DashboardFormsVentasPuntoVentaSAerializer(queryset, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)
